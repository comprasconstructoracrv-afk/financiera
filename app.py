from flask import Flask, render_template, request, redirect, session, flash, url_for, send_file
from models import db, Usuario, Credito, Cuota, Pago, ConfiguracionTasa, TasaPeriodo, Sede, TasaInteresVariable, InyeccionCapital, CambioTasaInteresCredito, AbonoCapital
from datetime import datetime, date, timedelta
import calendar
import os
from io import BytesIO
from math import floor
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import ImageReader


app = Flask(__name__)
app.secret_key = "supersecretkey"

database_url = os.getenv('DATABASE_URL', 'sqlite:///financiera.db')
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

from sqlalchemy import inspect

with app.app_context():
    inspector = inspect(db.engine)
    columnas_credito = [col['name'] for col in inspector.get_columns('credito')]

    if 'tipo_documento' not in columnas_credito:
        db.session.execute(
            db.text("ALTER TABLE credito ADD COLUMN tipo_documento VARCHAR(20) DEFAULT 'CC'")
        )
        db.session.commit()

    inspector = inspect(db.engine)

    columnas_abono = [col['name'] for col in inspector.get_columns('abono_capital')]

    if 'activo' not in columnas_abono:
        db.session.execute(
            db.text("ALTER TABLE abono_capital ADD COLUMN activo BOOLEAN DEFAULT TRUE")
        )

    if 'reversado' not in columnas_abono:
        db.session.execute(
            db.text("ALTER TABLE abono_capital ADD COLUMN reversado BOOLEAN DEFAULT FALSE")
        )

    if 'motivo_reversion' not in columnas_abono:
        db.session.execute(
            db.text("ALTER TABLE abono_capital ADD COLUMN motivo_reversion TEXT")
        )

    if 'fecha_reversion' not in columnas_abono:
        db.session.execute(
            db.text("ALTER TABLE abono_capital ADD COLUMN fecha_reversion TIMESTAMP")
        )

    db.session.execute(
        db.text("UPDATE abono_capital SET activo = TRUE WHERE activo IS NULL")
    )

    db.session.execute(
        db.text("UPDATE abono_capital SET reversado = FALSE WHERE reversado IS NULL")
    )

    db.session.commit()

@app.template_filter('cop')
def formato_cop(valor):
    try:
        if valor is None:
            return "$ 0"
        valor = int(round(float(valor)))
        return "$ {:,}".format(valor).replace(",", ".")
    except:
        return "$ 0"

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

def fecha_recibo_es(fecha):
    return f"{fecha.day:02d} {MESES_ES[fecha.month]} {fecha.year}"

UNIDADES = (
    '', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE'
)
DECENAS = (
    'DIEZ', 'ONCE', 'DOCE', 'TRECE', 'CATORCE', 'QUINCE',
    'DIECISEIS', 'DIECISIETE', 'DIECIOCHO', 'DIECINUEVE'
)
DIEZ_DIEZ = (
    '', '', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA',
    'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA'
)
CIENTOS = (
    '', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS',
    'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS'
)

MESES_ES_TEXTO = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

DIAS_ES = {
    0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
    4: "Viernes", 5: "Sábado", 6: "Domingo"
}

def fecha_larga_es(fecha):
    return f"{DIAS_ES[fecha.weekday()]}, {fecha.day:02d} de {MESES_ES_TEXTO[fecha.month]} del {fecha.year}"

def fecha_documento_es(fecha):
    return f"{fecha.day:02d} de {MESES_ES_TEXTO[fecha.month]} del {fecha.year}"

def convertir_menor_100(n):
    if n < 10:
        return UNIDADES[n]
    if 10 <= n < 20:
        return DECENAS[n - 10]
    if 20 <= n < 30:
        if n == 20:
            return 'VEINTE'
        return 'VEINTI' + UNIDADES[n - 20].lower()
    d = n // 10
    u = n % 10
    if u == 0:
        return DIEZ_DIEZ[d]
    return f"{DIEZ_DIEZ[d]} Y {UNIDADES[u]}"

def convertir_menor_1000(n):
    if n == 0:
        return ''
    if n == 100:
        return 'CIEN'
    c = n // 100
    resto = n % 100
    if c == 0:
        return convertir_menor_100(resto)
    if resto == 0:
        return CIENTOS[c]
    return f"{CIENTOS[c]} {convertir_menor_100(resto)}"

def numero_a_letras(n):
    n = int(round(n))
    if n == 0:
        return 'CERO PESOS'

    millones = n // 1000000
    miles = (n % 1000000) // 1000
    cientos = n % 1000

    partes = []

    if millones:
        if millones == 1:
            partes.append('UN MILLON')
        else:
            partes.append(f"{convertir_menor_1000(millones)} MILLONES")

    if miles:
        if miles == 1:
            partes.append('MIL')
        else:
            partes.append(f"{convertir_menor_1000(miles)} MIL")

    if cientos:
        partes.append(convertir_menor_1000(cientos))

    texto = ' '.join(partes).replace('  ', ' ').strip().upper()
    return f"{texto} PESOS"

def truncar_texto(canvas_obj, texto, x, y, ancho_max, font_name="Helvetica", font_size=10):
    texto = str(texto or "")
    while stringWidth(texto, font_name, font_size) > ancho_max and len(texto) > 0:
        texto = texto[:-1]
    canvas_obj.setFont(font_name, font_size)
    canvas_obj.drawString(x, y, texto)

def limpiar_valor_moneda(valor):
    if valor is None:
        return 0.0

    texto = str(valor).strip()
    texto = texto.replace('$', '').replace('.', '').replace(',', '').replace(' ', '')

    if texto == '':
        return 0.0

    return float(texto)


# 🔢 FUNCIÓN DE CÁLCULO
def calcular_cuota(monto, interes, cuotas):
    i = interes / 100
    cuota = monto * (i * (1 + i) ** cuotas) / ((1 + i) ** cuotas - 1)
    return round(cuota, 2)

def sumar_meses(fecha_base, meses, dia_fijo=None):
    """
    Mantiene el día original del crédito.
    Si el mes no tiene ese día (ej: febrero),
    usa el último día del mes.
    """

    if dia_fijo is None:
        dia_fijo = fecha_base.day

    mes = fecha_base.month - 1 + meses
    anio = fecha_base.year + mes // 12
    mes = mes % 12 + 1

    ultimo_dia_mes = calendar.monthrange(anio, mes)[1]
    dia = min(dia_fijo, ultimo_dia_mes)

    return datetime(anio, mes, dia)

def convertir_tasa_anual_a_mensual(tasa_anual):
    return round((((1 + (tasa_anual / 100)) ** (1/12)) - 1) * 100, 6)

def convertir_tasa_mensual_a_diaria(tasa_mensual):
    return round((tasa_mensual / 100) / 30, 10)

def obtener_o_crear_tasa_periodo(anio, mes, tasa_anual_base):
    tasa = TasaPeriodo.query.filter_by(anio=anio, mes=mes).first()

    if not tasa:
        tasa_mensual = convertir_tasa_anual_a_mensual(tasa_anual_base)
        tasa_diaria = convertir_tasa_mensual_a_diaria(tasa_mensual)

        tasa = TasaPeriodo(
            anio=anio,
            mes=mes,
            tasa_anual=tasa_anual_base,
            tasa_mensual=tasa_mensual,
            tasa_diaria=tasa_diaria
        )
        db.session.add(tasa)
        db.session.flush()

    return tasa


def generar_cuotas(credito_id, monto, interes, cuotas, fecha_base):
    saldo = round(monto, 2)
    tasa = interes / 100
    cuota_fija = calcular_cuota(monto, interes, cuotas)

    dia_original = fecha_base.day

    for n in range(cuotas):
        saldo_inicial = round(saldo, 2)
        interes_mes = round(saldo_inicial * tasa, 2)
        capital = round(cuota_fija - interes_mes, 2)
        saldo = round(saldo_inicial - capital, 2)

        if saldo < 0:
            capital = round(capital + saldo, 2)
            saldo = 0

        fecha_pago = sumar_meses(fecha_base, n, dia_fijo=dia_original)

        cuota = Cuota(
            credito_id=credito_id,
            numero=n + 1,
            fecha_pago=fecha_pago,
            saldo_inicial=saldo_inicial,
            valor_cuota=round(cuota_fija, 2),
            capital=capital,
            interes=interes_mes,
            saldo_restante=saldo,
            saldo_pendiente=round(cuota_fija, 2),
            dias_mora=0,
            interes_mora=0,
            total_cobro=round(cuota_fija, 2),
            estado='PENDIENTE'
        )

        db.session.add(cuota)

def ultimo_dia_mes(fecha):
    ultimo = calendar.monthrange(fecha.year, fecha.month)[1]
    return date(fecha.year, fecha.month, ultimo)

def obtener_tasa_periodo(anio, mes):
    tasa_periodo = TasaPeriodo.query.filter_by(anio=anio, mes=mes).first()

    if tasa_periodo:
        return tasa_periodo

    return ConfiguracionTasa.query.first()

def obtener_tasa_interes_variable(anio, tasa_inicial):
    tasa = TasaInteresVariable.query.filter_by(anio=anio).first()

    if tasa:
        return tasa.tasa_mensual

    return tasa_inicial

def actualizar_mora_credito(credito, fecha_corte=None):
    if isinstance(credito, int):
        credito = Credito.query.get_or_404(credito)

    if fecha_corte is None:
        fecha_corte = date.today()

    cuotas_credito = Cuota.query.filter_by(
        credito_id=credito.id
    ).order_by(Cuota.numero).all()

    for cuota in cuotas_credito:
        fecha_vencimiento = cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago

        if cuota.estado == 'LIQUIDADA':
            cuota.saldo_pendiente = 0
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = 0
            continue

        pagos_activos = Pago.query.filter(
            Pago.cuota_id == cuota.id,
            Pago.activo == True,
            Pago.tipo_pago != 'ABONO_CAPITAL'
        ).all()

        total_pagado_activo = round(sum(p.valor or 0 for p in pagos_activos), 2)

        valor_cuota = round(cuota.valor_cuota or 0, 2)

        total_aplicado_cuota = round(sum(
            (p.valor_aplicado_interes or 0) + (p.valor_aplicado_capital or 0)
            for p in pagos_activos
        ), 2)

        total_mora_generada_historica = round(max(
            [p.mora_generada_al_pago or 0 for p in pagos_activos] or [0]
        ), 2)

        total_mora_pagada = round(sum(
            p.valor_aplicado_mora or 0
            for p in pagos_activos
        ), 2)

        mora_pendiente_historica = round(
            total_mora_generada_historica - total_mora_pagada,
            2
        )

        if mora_pendiente_historica < 0:
            mora_pendiente_historica = 0

        saldo_base = round(valor_cuota - total_aplicado_cuota, 2)

        if saldo_base <= 1:
            saldo_base = 0

        tasa = obtener_tasa_periodo(fecha_vencimiento.year, fecha_vencimiento.month)
        tasa_mensual = tasa.tasa_mensual if tasa else 0
        tasa_diaria = tasa.tasa_diaria if tasa else 0

        cuota.tasa_mora_mensual_cuota = tasa_mensual
        cuota.porcentaje_mora_aplicado = tasa_mensual

        if saldo_base <= 0:
            cuota.saldo_pendiente = 0

            if mora_pendiente_historica > 1:
                cuota.interes_mora = round(mora_pendiente_historica, 2)
                cuota.total_cobro = round(mora_pendiente_historica, 2)
                cuota.estado = 'EN MORA'

                pagos_con_mora = [
                    p for p in pagos_activos
                    if (p.mora_generada_al_pago or 0) > 0
                ]

                if pagos_con_mora:
                    cuota.dias_mora = max(
                        p.dias_mora_pagados or 0
                        for p in pagos_con_mora
                    )
                elif fecha_corte > fecha_vencimiento:
                    cuota.dias_mora = (fecha_corte - fecha_vencimiento).days
                else:
                    cuota.dias_mora = 0

            else:
                cuota.dias_mora = 0
                cuota.interes_mora = 0
                cuota.total_cobro = 0
                cuota.estado = 'PAGADA'

            continue

        cuota.saldo_pendiente = saldo_base

        if fecha_corte > fecha_vencimiento:
            dias_mora = (fecha_corte - fecha_vencimiento).days
            interes_mora = saldo_base * tasa_diaria * dias_mora

            cuota.dias_mora = dias_mora
            cuota.interes_mora = round(interes_mora + mora_pendiente_historica, 2)
            cuota.total_cobro = round(saldo_base + cuota.interes_mora, 2)
            cuota.estado = 'EN MORA'
        else:
            cuota.dias_mora = 0
            cuota.interes_mora = round(mora_pendiente_historica, 2)
            cuota.total_cobro = round(saldo_base + cuota.interes_mora, 2)

            if cuota.interes_mora > 1:
                cuota.estado = 'EN MORA'
            elif total_pagado_activo > 0:
                cuota.estado = 'ABONO'
            else:
                cuota.estado = 'PENDIENTE'

def recalcular_cuotas_pendientes(credito, cuota_actual_numero, fecha_base):
    cuotas_futuras = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.numero > cuota_actual_numero
    ).order_by(Cuota.numero).all()

    if not cuotas_futuras:
        return

    def obtener_prepago_pago(pago):
        prepago = pago.valor_aplicado_prepago_capital or 0

        if prepago and prepago > 0:
            return round(prepago, 2)

        valor_pago = pago.valor or 0
        aplicado_interes = pago.valor_aplicado_interes or 0
        aplicado_capital = pago.valor_aplicado_capital or 0
        aplicado_mora = pago.valor_aplicado_mora or 0

        excedente = (
            valor_pago
            - aplicado_interes
            - aplicado_capital
            - aplicado_mora
        )

        if excedente > 0:
            return round(excedente, 2)

        return 0

    cuota_anterior = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.numero == cuota_actual_numero
    ).first()

    if cuota_anterior:
        saldo = round(cuota_anterior.saldo_restante or 0, 2)

        pagos_anterior = Pago.query.filter(
            Pago.cuota_id == cuota_anterior.id,
            Pago.activo == True,
            Pago.tipo_pago == 'PAGO_CUOTA'
        ).all()

        prepago_anterior = round(sum(
            obtener_prepago_pago(p)
            for p in pagos_anterior
        ), 2)

        saldo = round(saldo - prepago_anterior, 2)

        fecha_base_recalculo = (
            cuota_anterior.fecha_pago.date()
            if isinstance(cuota_anterior.fecha_pago, datetime)
            else cuota_anterior.fecha_pago
        )
    else:
        saldo = round(credito.monto_financiado or 0, 2)

        fecha_base_recalculo = (
            credito.fecha_creacion.date()
            if isinstance(credito.fecha_creacion, datetime)
            else credito.fecha_creacion
        )

    if saldo < 0:
        saldo = 0

    abonos_credito = AbonoCapital.query.filter_by(
        credito_id=credito.id,
        activo=True
    ).order_by(
        AbonoCapital.fecha.asc(),
        AbonoCapital.id.asc()
    ).all()

    abonos_aplicados = set()

    for abono in abonos_credito:
        fecha_abono = (
            abono.fecha.date()
            if isinstance(abono.fecha, datetime)
            else abono.fecha
        )

        if fecha_abono <= fecha_base_recalculo:
            abonos_aplicados.add(abono.id)

    tasa_credito = (credito.interes or 0) / 100

    config_tasa = ConfiguracionTasa.query.filter_by(
        nombre='TASA_MORA'
    ).first()

    fecha_creacion_credito = (
        credito.fecha_creacion.date()
        if isinstance(credito.fecha_creacion, datetime)
        else credito.fecha_creacion
    )

    dia_original = fecha_creacion_credito.day

    for cuota in cuotas_futuras:
        nueva_fecha = sumar_meses(
            fecha_creacion_credito,
            cuota.numero - 1,
            dia_fijo=dia_original
        )

        nueva_fecha_date = (
            nueva_fecha.date()
            if isinstance(nueva_fecha, datetime)
            else nueva_fecha
        )

        total_abonos_periodo = 0

        for abono in abonos_credito:
            fecha_abono = (
                abono.fecha.date()
                if isinstance(abono.fecha, datetime)
                else abono.fecha
            )

            if (
                abono.id not in abonos_aplicados
                and fecha_abono <= nueva_fecha_date
            ):
                total_abonos_periodo += round(abono.valor or 0, 2)
                abonos_aplicados.add(abono.id)

        saldo = round(saldo - total_abonos_periodo, 2)

        if saldo < 0:
            saldo = 0

        cuotas_restantes = max(
            credito.cuotas - cuota.numero + 1,
            1
        )

        valor_cuota = round(
            calcular_cuota(
                saldo,
                credito.interes,
                cuotas_restantes
            ),
            2
        )

        saldo_inicial = round(saldo, 2)

        interes_mes = round(
            saldo_inicial * tasa_credito,
            2
        )

        capital = round(
            valor_cuota - interes_mes,
            2
        )

        saldo_restante = round(
            saldo_inicial - capital,
            2
        )

        if saldo_restante < 0:
            capital = round(
                capital + saldo_restante,
                2
            )

            saldo_restante = 0

            valor_cuota = round(
                capital + interes_mes,
                2
            )

        pagos_cuota = Pago.query.filter(
            Pago.cuota_id == cuota.id,
            Pago.activo == True,
            Pago.tipo_pago == 'PAGO_CUOTA'
        ).all()

        prepago_cuota = round(sum(
            obtener_prepago_pago(p)
            for p in pagos_cuota
        ), 2)

        saldo_despues_prepago = round(
            saldo_restante - prepago_cuota,
            2
        )

        if saldo_despues_prepago < 0:
            saldo_despues_prepago = 0

        tasa_periodo = TasaPeriodo.query.filter_by(
            anio=nueva_fecha.year,
            mes=nueva_fecha.month
        ).first()

        cuota.fecha_pago = nueva_fecha
        cuota.saldo_inicial = saldo_inicial
        cuota.valor_cuota = valor_cuota
        cuota.capital = capital
        cuota.interes = interes_mes
        cuota.saldo_restante = saldo_restante

        if pagos_cuota:
            cuota.saldo_pendiente = 0
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = 0
            cuota.estado = 'PAGADA'
        else:
            cuota.saldo_pendiente = valor_cuota
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = valor_cuota
            cuota.estado = 'PENDIENTE'

        if tasa_periodo:
            cuota.tasa_mora_mensual_cuota = (
                tasa_periodo.tasa_mensual
            )
        elif config_tasa:
            cuota.tasa_mora_mensual_cuota = (
                config_tasa.tasa_mensual
            )
        else:
            cuota.tasa_mora_mensual_cuota = 0

        cuota.porcentaje_mora_aplicado = (
            cuota.tasa_mora_mensual_cuota
        )

        saldo = saldo_despues_prepago

    credito.saldo_actual = round(saldo, 2)

def obtener_tasa_credito_en_cuota(credito, numero_cuota, fecha_pago):
    cambio = CambioTasaInteresCredito.query.filter(
        CambioTasaInteresCredito.credito_id == credito.id,
        CambioTasaInteresCredito.numero_cuota <= numero_cuota
    ).order_by(CambioTasaInteresCredito.numero_cuota.desc()).first()

    if cambio:
        return cambio.tasa_nueva

    return credito.interes or 0

def recalcular_cuotas_variables_pendientes(credito, cuota_actual_numero, fecha_base):
    if not (credito.tipo_cuota == 'VARIABLE' and credito.tipo_interes == 'VARIABLE'):
        return

    cuotas_futuras = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.numero > cuota_actual_numero
    ).order_by(Cuota.numero.asc()).all()

    if not cuotas_futuras:
        return

    cuota_anterior = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.numero == cuota_actual_numero
    ).first()

    if cuota_anterior:
        saldo = round(cuota_anterior.saldo_restante or 0, 2)

        pagos_cuota_anterior = Pago.query.filter(
            Pago.cuota_id == cuota_anterior.id,
            Pago.activo == True,
            Pago.tipo_pago == 'PAGO_CUOTA'
        ).all()

        total_prepago_anterior = round(sum(
            p.valor_aplicado_prepago_capital or 0
            for p in pagos_cuota_anterior
        ), 2)

        saldo = round(saldo - total_prepago_anterior, 2)

        fecha_base_recalculo = (
            cuota_anterior.fecha_pago.date()
            if isinstance(cuota_anterior.fecha_pago, datetime)
            else cuota_anterior.fecha_pago
        )
    else:
        saldo = round(credito.monto_financiado or 0, 2)

        fecha_base_recalculo = (
            credito.fecha_creacion.date()
            if isinstance(credito.fecha_creacion, datetime)
            else credito.fecha_creacion
        )

    dia_original = credito.fecha_creacion.day

    abonos_credito = AbonoCapital.query.filter_by(
        credito_id=credito.id,
        activo=True
    ).order_by(AbonoCapital.fecha.asc(), AbonoCapital.id.asc()).all()

    abonos_aplicados = set()

    # Los abonos anteriores o iguales a la cuota base se consideran ya incluidos
    for abono in abonos_credito:
        fecha_abono = abono.fecha.date() if isinstance(abono.fecha, datetime) else abono.fecha

        if fecha_abono <= fecha_base_recalculo:
            abonos_aplicados.add(abono.id)

    for cuota in cuotas_futuras:
        fecha_pago = sumar_meses(
            credito.fecha_creacion,
            cuota.numero - 1,
            dia_fijo=dia_original
        )

        fecha_pago_date = fecha_pago.date() if isinstance(fecha_pago, datetime) else fecha_pago

        total_abonos_periodo = 0

        for abono in abonos_credito:
            fecha_abono = abono.fecha.date() if isinstance(abono.fecha, datetime) else abono.fecha

            if abono.id not in abonos_aplicados and fecha_abono <= fecha_pago_date:
                total_abonos_periodo += round(abono.valor or 0, 2)
                abonos_aplicados.add(abono.id)

        inyecciones_cuota = InyeccionCapital.query.filter_by(
            credito_id=credito.id,
            numero_cuota=cuota.numero
        ).all()

        adicion_capital = round(sum(
            i.valor or 0
            for i in inyecciones_cuota
        ), 2)

        saldo = round(saldo - total_abonos_periodo + adicion_capital, 2)

        if saldo < 0:
            saldo = 0

        cuotas_restantes = credito.cuotas - cuota.numero + 1

        tasa_mes = obtener_tasa_credito_en_cuota(
            credito,
            cuota.numero,
            fecha_pago
        )

        saldo_inicial = round(saldo, 2)
        interes_mes = round(saldo_inicial * (tasa_mes / 100), 2)
        capital = round(saldo_inicial / max(cuotas_restantes, 1), 2)
        valor_cuota = round(capital + interes_mes, 2)

        saldo = round(saldo_inicial - capital, 2)

        if saldo < 0:
            capital = round(capital + saldo, 2)
            saldo = 0
            valor_cuota = round(capital + interes_mes, 2)

        cuota.fecha_pago = fecha_pago
        cuota.saldo_inicial = saldo_inicial
        cuota.valor_cuota = valor_cuota
        cuota.capital = capital
        cuota.interes = interes_mes
        cuota.saldo_restante = saldo
        cuota.saldo_pendiente = valor_cuota
        cuota.dias_mora = 0
        cuota.interes_mora = 0
        cuota.total_cobro = valor_cuota
        cuota.estado = 'PENDIENTE'

    credito.saldo_actual = round(saldo, 2)

def aplicar_pago_deuda_fecha(credito, fecha_pago, valor_pago, medio_pago):
    actualizar_mora_credito(credito, fecha_pago)

    cuotas_exigibles = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    cuotas_exigibles = [
        cuota for cuota in cuotas_exigibles
        if (cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago) <= fecha_pago
    ]

    restante = round(valor_pago, 2)
    hubo_abono_extra_capital = False
    ultima_cuota_tocada = None
    pagos_creados_ids = []

    for cuota in cuotas_exigibles:
        if restante <= 0:
            break

        ultima_cuota_tocada = cuota

        dias_mora_al_pago = cuota.dias_mora or 0
        mora_generada_al_pago = round(cuota.interes_mora or 0, 2)
        saldo_pendiente_antes_pago = round(cuota.saldo_pendiente or 0, 2)
        total_exigible_al_pago = round((cuota.saldo_pendiente or 0) + (cuota.interes_mora or 0), 2)

        pago = Pago(
            cuota_id=cuota.id,
            fecha=datetime.combine(fecha_pago, datetime.min.time()),
            valor=0,
            medio_pago=medio_pago,
            tipo_pago='PAGO_DEUDA_FECHA',
            dias_mora_pagados=dias_mora_al_pago,
            mora_generada_al_pago=mora_generada_al_pago,
            saldo_pendiente_antes_pago=saldo_pendiente_antes_pago,
            total_exigible_al_pago=total_exigible_al_pago
        )

        valor_aplicado_cuota = 0
        valor_aplicado_mora = 0
        valor_aplicado_interes = 0
        valor_aplicado_capital = 0

        # 1. Cubrir cuota primero
        if cuota.saldo_pendiente > 0 and restante > 0:
            aplicado_cuota = min(restante, round(cuota.saldo_pendiente, 2))
            cuota.saldo_pendiente = round(cuota.saldo_pendiente - aplicado_cuota, 2)
            restante = round(restante - aplicado_cuota, 2)
            valor_aplicado_cuota = round(aplicado_cuota, 2)

            interes_cuota = round(cuota.interes or 0, 2)
            valor_aplicado_interes = min(valor_aplicado_cuota, interes_cuota)
            valor_aplicado_capital = round(max(valor_aplicado_cuota - valor_aplicado_interes, 0), 2)

            if cuota.saldo_pendiente <= 0:
                cuota.saldo_pendiente = 0
                credito.saldo_actual = round(cuota.saldo_restante, 2)

        # 2. Luego cubrir mora
        if cuota.interes_mora > 0 and restante > 0:
            aplicado_mora = min(restante, round(cuota.interes_mora, 2))
            cuota.interes_mora = round(cuota.interes_mora - aplicado_mora, 2)
            restante = round(restante - aplicado_mora, 2)
            valor_aplicado_mora = aplicado_mora

        pago.valor = round(valor_aplicado_cuota + valor_aplicado_mora, 2)
        pago.valor_aplicado_interes = round(valor_aplicado_interes, 2)
        pago.valor_aplicado_capital = round(valor_aplicado_capital, 2)
        pago.valor_aplicado_mora = round(valor_aplicado_mora, 2)
        if pago.valor > 0:
            db.session.add(pago)
            db.session.flush()
            pagos_creados_ids.append(pago.id)

        cuota.saldo_pendiente = round(max(cuota.saldo_pendiente, 0), 2)
        cuota.interes_mora = round(max(cuota.interes_mora, 0), 2)

        if cuota.saldo_pendiente <= 0 and cuota.interes_mora <= 0:
            cuota.saldo_pendiente = 0
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = 0
            cuota.estado = 'PAGADA'
        elif cuota.saldo_pendiente <= 0 and cuota.interes_mora > 0:
            cuota.saldo_pendiente = 0
            cuota.total_cobro = round(cuota.interes_mora, 2)
            cuota.estado = 'ABONO'
        else:
            cuota.total_cobro = round(cuota.saldo_pendiente + cuota.interes_mora, 2)
            if cuota.dias_mora > 0:
                cuota.estado = 'EN MORA'
            else:
                cuota.estado = 'ABONO'

    # Solo si ya cubrió toda la deuda exigible a hoy y sobra dinero, va a capital
    if restante > 0:
        credito.saldo_actual = round(credito.saldo_actual - restante, 2)
        if credito.saldo_actual < 0:
            credito.saldo_actual = 0
        hubo_abono_extra_capital = True

    if hubo_abono_extra_capital and ultima_cuota_tocada:
        recalcular_cuotas_pendientes(
            credito=credito,
            cuota_actual_numero=ultima_cuota_tocada.numero,
            fecha_base=ultima_cuota_tocada.fecha_pago
        )

        db.session.commit()
        return pagos_creados_ids

def calcular_componentes_liquidacion(credito, fecha_corte):
    cuotas_activas = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    if not cuotas_activas:
        return {
            'cuota_actual': None,
            'capital_insoluto': 0,
            'interes_corriente': 0,
            'total_mora': 0,
            'total_liquidacion': 0
        }

    cuotas_vencidas_o_del_mes = []

    for cuota in cuotas_activas:
        fecha_vencimiento = (
            cuota.fecha_pago.date()
            if isinstance(cuota.fecha_pago, datetime)
            else cuota.fecha_pago
        )

        if fecha_vencimiento <= fecha_corte:
            cuotas_vencidas_o_del_mes.append(cuota)

    if cuotas_vencidas_o_del_mes:
        cuota_actual = cuotas_vencidas_o_del_mes[0]
        cuotas_a_cobrar = cuotas_vencidas_o_del_mes
    else:
        cuota_actual = cuotas_activas[0]
        cuotas_a_cobrar = [cuota_actual]

    capital_insoluto = round(cuota_actual.saldo_inicial or 0, 2)

    interes_corriente = round(sum(
        c.interes or 0
        for c in cuotas_a_cobrar
    ), 2)

    total_mora = round(sum(
        c.interes_mora or 0
        for c in cuotas_a_cobrar
    ), 2)

    total_liquidacion = round(
        capital_insoluto + interes_corriente + total_mora,
        2
    )

    return {
        'cuota_actual': cuota_actual,
        'capital_insoluto': capital_insoluto,
        'interes_corriente': interes_corriente,
        'total_mora': total_mora,
        'total_liquidacion': total_liquidacion
    }

def generar_cuotas(credito_id, monto, interes, cuotas, fecha_base):
    saldo = round(monto, 2)
    tasa_credito = interes / 100
    cuota_fija = calcular_cuota(monto, interes, cuotas)

    config_tasa = ConfiguracionTasa.query.filter_by(nombre='TASA_MORA').first()

    for n in range(cuotas):
        saldo_inicial = round(saldo, 2)
        interes_mes = round(saldo_inicial * tasa_credito, 2)
        capital = round(cuota_fija - interes_mes, 2)
        saldo = round(saldo_inicial - capital, 2)

        if saldo < 0:
            saldo = 0

        fecha_pago = sumar_meses(fecha_base, n)

        tasa_periodo = TasaPeriodo.query.filter_by(
            anio=fecha_pago.year,
            mes=fecha_pago.month,
          
        ).first()

        nueva_cuota = Cuota(
            credito_id=credito_id,
            numero=n + 1,
            fecha_pago=fecha_pago,
            valor_cuota=cuota_fija,
            saldo_inicial=saldo_inicial,
            capital=capital,
            interes=interes_mes,
            saldo_restante=saldo,
            saldo_pendiente=cuota_fija,
            tasa_mora_mensual_cuota=tasa_periodo.tasa_mensual if tasa_periodo else config_tasa.tasa_mensual,
            dias_mora=0,
            interes_mora=0,
            total_cobro=cuota_fija,
            estado='PENDIENTE'
        )
        db.session.add(nueva_cuota)

def generar_cuotas_variables(credito_id, monto, interes_inicial, cuotas, fecha_base):
    saldo = round(monto, 2)
    dia_original = fecha_base.day

    for n in range(cuotas):
        numero_cuota = n + 1
        fecha_pago = sumar_meses(fecha_base, n, dia_fijo=dia_original)

        credito = Credito.query.get(credito_id)

        tasa_mes = obtener_tasa_interes_variable(
            fecha_pago.year,
            interes_inicial
        )

        inyecciones = InyeccionCapital.query.filter_by(
            credito_id=credito_id,
            numero_cuota=numero_cuota
        ).all()

        adicion_capital = round(sum(i.valor or 0 for i in inyecciones), 2)

        saldo_inicial = round(saldo + adicion_capital, 2)

        interes_mes = round(saldo_inicial * (tasa_mes / 100), 2)

        capital_base = round(saldo_inicial / max(cuotas - n, 1), 2)

        valor_cuota = round(capital_base + interes_mes, 2)

        saldo = round(saldo_inicial - capital_base, 2)

        if saldo < 0:
            capital_base = round(capital_base + saldo, 2)
            saldo = 0

        nueva_cuota = Cuota(
            credito_id=credito_id,
            numero=numero_cuota,
            fecha_pago=fecha_pago,
            valor_cuota=valor_cuota,
            saldo_inicial=saldo_inicial,
            capital=capital_base,
            interes=interes_mes,
            saldo_restante=saldo,
            saldo_pendiente=valor_cuota,
            tasa_mora_mensual_cuota=0,
            porcentaje_mora_aplicado=0,
            dias_mora=0,
            interes_mora=0,
            total_cobro=valor_cuota,
            estado='PENDIENTE'
        )

        db.session.add(nueva_cuota)

    if credito:
        credito.cuota_mensual = 0
        credito.saldo_actual = round(saldo, 2)

def construir_datos_reporte(anio_seleccionado, sede_seleccionada):
    ...
    return {
        'resumen_general': resumen_general,
        'resumen_por_sede': resumen_por_sede,
        'resumen_mensual': resumen_mensual,
        'labels_sedes': labels_sedes,
        'saldo_actual_sedes': saldo_actual_sedes,
        'interes_causado_sedes': interes_causado_sedes,
        'interes_recaudado_sedes': interes_recaudado_sedes,
        'mora_causada_sedes': mora_causada_sedes,
        'mora_recaudada_sedes': mora_recaudada_sedes,
        'labels_meses': labels_meses,
        'interes_causado_meses': interes_causado_meses,
        'interes_recaudado_meses': interes_recaudado_meses,
        'mora_causada_meses': mora_causada_meses,
        'mora_recaudada_meses': mora_recaudada_meses,
    }

def numero_a_letras_es(n):
    unidades = (
        '', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE'
    )
    decenas = (
        '', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA',
        'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA'
    )
    especiales = {
        11: 'ONCE', 12: 'DOCE', 13: 'TRECE', 14: 'CATORCE', 15: 'QUINCE',
        16: 'DIECISEIS', 17: 'DIECISIETE', 18: 'DIECIOCHO', 19: 'DIECINUEVE',
        21: 'VEINTIUNO', 22: 'VEINTIDOS', 23: 'VEINTITRES', 24: 'VEINTICUATRO',
        25: 'VEINTICINCO', 26: 'VEINTISEIS', 27: 'VEINTISIETE',
        28: 'VEINTIOCHO', 29: 'VEINTINUEVE'
    }
    centenas = (
        '', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS',
        'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS'
    )

    def convertir_menor_100(num):
        if num < 10:
            return unidades[num]
        if num in especiales:
            return especiales[num]
        if num == 10:
            return 'DIEZ'
        if num < 20:
            return 'DIECI' + unidades[num - 10].lower().upper()
        if num == 20:
            return 'VEINTE'
        if num < 30:
            return especiales.get(num, 'VEINTI' + unidades[num - 20].lower().upper())
        d = num // 10
        u = num % 10
        if u == 0:
            return decenas[d]
        return f"{decenas[d]} Y {unidades[u]}"

    def convertir_menor_1000(num):
        if num == 0:
            return ''
        if num == 100:
            return 'CIEN'
        c = num // 100
        resto = num % 100
        if c == 0:
            return convertir_menor_100(resto)
        if resto == 0:
            return centenas[c]
        return f"{centenas[c]} {convertir_menor_100(resto)}"

    def convertir(num):
        if num == 0:
            return 'CERO'
        if num < 1000:
            return convertir_menor_1000(num)
        if num < 1000000:
            miles = num // 1000
            resto = num % 1000
            if miles == 1:
                texto_miles = 'MIL'
            else:
                texto_miles = f"{convertir_menor_1000(miles)} MIL"
            if resto == 0:
                return texto_miles
            return f"{texto_miles} {convertir_menor_1000(resto)}"
        millones = num // 1000000
        resto = num % 1000000
        if millones == 1:
            texto_millones = 'UN MILLON'
        else:
            texto_millones = f"{convertir(millones)} MILLONES"
        if resto == 0:
            return texto_millones
        return f"{texto_millones} {convertir(resto)}"

    entero = int(round(float(n or 0)))
    return convertir(entero)


def formatear_fecha_larga(fecha_obj):
    if not fecha_obj:
        return ''
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return f"{fecha_obj.day} días del mes de {meses[fecha_obj.month]} del {fecha_obj.year}"


def fecha_a_ddmmyyyy(fecha_obj):
    if not fecha_obj:
        return ''
    return fecha_obj.strftime('%d/%m/%Y')


def obtener_fecha_credito_real(credito):
    if hasattr(credito, 'fecha_credito') and credito.fecha_credito:
        return credito.fecha_credito.date() if isinstance(credito.fecha_credito, datetime) else credito.fecha_credito
    if hasattr(credito, 'fecha_creacion') and credito.fecha_creacion:
        return credito.fecha_creacion.date() if isinstance(credito.fecha_creacion, datetime) else credito.fecha_creacion
    return date.today()


def obtener_primera_y_ultima_cuota(credito):
    cuotas = []
    if hasattr(credito, 'cuotas_rel') and credito.cuotas_rel:
        cuotas = sorted(credito.cuotas_rel, key=lambda c: c.numero)

    if not cuotas:
        return None, None, None

    primera = cuotas[0]
    ultima = cuotas[-1]

    fecha_inicio = primera.fecha_pago.date() if isinstance(primera.fecha_pago, datetime) else primera.fecha_pago
    fecha_final = ultima.fecha_pago.date() if isinstance(ultima.fecha_pago, datetime) else ultima.fecha_pago
    dia_pago = fecha_inicio.day if fecha_inicio else ''

    return fecha_inicio, fecha_final, dia_pago

def inicializar_sedes():
    sedes_base = ['IBAGUE', 'ESPINAL', 'GIRARDOT', 'CRV']

    for nombre in sedes_base:
        existe = Sede.query.filter_by(nombre=nombre).first()
        if not existe:
            db.session.add(Sede(nombre=nombre, activa=True))

    db.session.commit()


# 🧱 CREAR BD + USUARIO ADMIN

with app.app_context():
    db.create_all()
    inicializar_sedes()

with app.app_context():
    db.create_all()

    if not Usuario.query.filter_by(username='admin').first():
        nuevo = Usuario(username='admin', password='1234', rol='admin')
        db.session.add(nuevo)
        db.session.commit()

    if not ConfiguracionTasa.query.filter_by(nombre='TASA_MORA').first():
        tasa_anual = 25.52
        tasa_mensual = convertir_tasa_anual_a_mensual(tasa_anual)
        tasa_diaria = convertir_tasa_mensual_a_diaria(tasa_mensual)

        config = ConfiguracionTasa(
            nombre='TASA_MORA',
            tasa_anual=tasa_anual,
            tasa_mensual=tasa_mensual,
            tasa_diaria=tasa_diaria
        )
        db.session.add(config)
        db.session.commit()

# 🔐 LOGIN
@app.route('/')
def inicio():
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():

    # LIMPIAR MENSAJES FLASH ACUMULADOS
    session.pop('_flashes', None)

    if request.method == 'POST':
        user = request.form['username']
        password = request.form['password']

        usuario = Usuario.query.filter_by(
            username=user,
            password=password
        ).first()

        if usuario:
            session['user'] = usuario.username
            session['rol'] = usuario.rol
            return redirect('/dashboard')

        else:
            flash("Usuario o contraseña incorrectos", "error")
            return redirect('/login')

    return render_template('login.html')

# 📊 DASHBOARD

@app.route('/crear_credito', methods=['GET', 'POST'])
def crear_credito():
    if 'user' not in session:
        return redirect('/login')

    sedes = Sede.query.filter_by(activa=True).order_by(Sede.nombre.asc()).all()

    if request.method == 'POST':
        try:
            cliente = request.form['cliente'].strip()
            tipo_documento = request.form.get('tipo_documento', 'CC').strip()
            cedula_cliente = request.form.get('cedula_cliente', '').strip()
            telefono_1 = request.form.get('telefono_1', '').strip()
            telefono_2 = request.form.get('telefono_2', '').strip()
            direccion_cliente = request.form.get('direccion_cliente', '').strip()
            correo_cliente = request.form.get('correo_cliente', '').strip()

            numero_pagare = request.form['numero_pagare'].strip()
            sede = request.form.get('sede', 'CRV')
            monto = limpiar_valor_moneda(request.form['monto'])
            interes = float(request.form['interes'])
            cuotas = int(request.form['cuotas'])
            fecha_credito = datetime.strptime(request.form['fecha_credito'], '%Y-%m-%d')

            tipo_cuota = request.form.get('tipo_cuota', 'FIJA')
            tipo_interes = request.form.get('tipo_interes', 'FIJO')
            permite_inyeccion = request.form.get('permite_inyeccion_capital', '').strip().upper() in ['SI', 'SÍ']

            tiene_codeudor = request.form.get('tiene_codeudor') == 'SI'

            codeudor_nombre = request.form.get('codeudor_nombre', '').strip() if tiene_codeudor else None
            codeudor_identificacion = request.form.get('codeudor_identificacion', '').strip() if tiene_codeudor else None
            codeudor_direccion = request.form.get('codeudor_direccion', '').strip() if tiene_codeudor else None
            codeudor_telefono = request.form.get('codeudor_telefono', '').strip() if tiene_codeudor else None
            codeudor_correo = request.form.get('codeudor_correo', '').strip() if tiene_codeudor else None

            abono_inicial_texto = request.form.get('abono_inicial', '').strip()
            abono_inicial = limpiar_valor_moneda(abono_inicial_texto) if abono_inicial_texto else 0

            monto_financiado = monto - abono_inicial

            if monto_financiado <= 0:
                flash("El monto financiado debe ser mayor que cero", "error")
                return render_template('crear_credito.html', sedes=sedes)

            cuota = calcular_cuota(monto_financiado, interes, cuotas)

            config_tasa = ConfiguracionTasa.query.filter_by(nombre='TASA_MORA').first()

            nuevo = Credito(
                numero_pagare=numero_pagare,
                cliente=cliente,
                sede=sede,
                tipo_documento=tipo_documento,
                cedula_cliente=cedula_cliente,
                telefono_1=telefono_1,
                telefono_2=telefono_2,
                direccion_cliente=direccion_cliente,
                correo_cliente=correo_cliente,
                tiene_codeudor=tiene_codeudor,
                codeudor_nombre=codeudor_nombre,
                codeudor_identificacion=codeudor_identificacion,
                codeudor_direccion=codeudor_direccion,
                codeudor_telefono=codeudor_telefono,
                codeudor_correo=codeudor_correo,
                monto=monto,
                abono_inicial=abono_inicial,
                monto_financiado=monto_financiado,
                saldo_actual=monto_financiado,
                interes=interes,
                cuotas=cuotas,
                cuota_mensual=cuota,
                tasa_mora_anual=config_tasa.tasa_anual,
                tasa_mora_mensual=config_tasa.tasa_mensual,
                tasa_mora_diaria=config_tasa.tasa_diaria,
                fecha_creacion=fecha_credito,
                tipo_cuota=tipo_cuota,
                tipo_interes=tipo_interes,
                permite_inyeccion_capital=permite_inyeccion
            )

            db.session.add(nuevo)
            db.session.commit()

            if tipo_cuota == 'VARIABLE':
                generar_cuotas_variables(
                    nuevo.id,
                    monto_financiado,
                    interes,
                    cuotas,
                    fecha_credito
                )
            else:
                generar_cuotas(
                    nuevo.id,
                    monto_financiado,
                    interes,
                    cuotas,
                    fecha_credito
                )
            db.session.commit()

            flash("Crédito creado correctamente", "success")
            return redirect(url_for('pagare_credito', credito_id=nuevo.id, desde='crear'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar el crédito: {str(e)}", "error")
            return render_template('crear_credito.html', sedes=sedes)

    return render_template('crear_credito.html', sedes=sedes)

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect('/login')

    sedes_db = Sede.query.filter_by(activa=True).order_by(Sede.nombre.asc()).all()
    resumen_sedes = []

    for sede_obj in sedes_db:
        sede = sede_obj.nombre

        creditos = Credito.query.filter_by(sede=sede).all()

        total = len(creditos)
        en_mora = 0
        cancelados = 0
        al_dia = 0

        for credito in creditos:
            cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

            if not cuotas:
                continue

            if any(c.estado == 'EN MORA' for c in cuotas):
                en_mora += 1
            elif all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
                cancelados += 1
            else:
                al_dia += 1

        resumen_sedes.append({
            'sede': sede,
            'total': total,
            'en_mora': en_mora,
            'cancelados': cancelados,
            'al_dia': al_dia
        })

    return render_template(
        'dashboard.html',
        resumen_sedes=resumen_sedes
    )

@app.route('/creditos/<sede>')
def creditos_sede(sede):
    if 'user' not in session:
        return redirect('/login')

    sede = sede.strip().upper()

    return render_template(
        'creditos_sede.html',
        sede_actual=sede
    )

# 🚪 LOGOUT
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/ver_creditos/<sede>')
def ver_creditos(sede):
    if 'user' not in session:
        return redirect('/login')

    sede = sede.strip().upper()
    hoy = date.today()

    creditos = Credito.query.filter_by(
        sede=sede
    ).order_by(Credito.fecha_creacion.desc()).all()

    resumen_creditos = []

    for credito in creditos:
        actualizar_mora_credito(credito, hoy)

        cuotas = Cuota.query.filter_by(
            credito_id=credito.id
        ).all()

        saldo_actual_credito = sum(
            (c.saldo_pendiente or 0) + (c.interes_mora or 0)
            for c in cuotas
            if c.estado not in ['PAGADA', 'LIQUIDADA']
        )

        if not cuotas:
            continue

        if any(c.estado == 'EN MORA' for c in cuotas):
            estado_credito = 'EN MORA'
        elif all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
            estado_credito = 'CANCELADO'
        elif any(c.estado == 'ABONO' for c in cuotas):
            estado_credito = 'CON ABONOS'
        else:
            estado_credito = 'AL DÍA'

        if all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
            continue

        total_inyecciones = db.session.query(
            db.func.coalesce(
                db.func.sum(InyeccionCapital.valor),
                0
            )
        ).filter(
            InyeccionCapital.credito_id == credito.id
        ).scalar()

        credito.monto_total_con_inyecciones = (
            (credito.monto or 0) + (total_inyecciones or 0)
        )

        resumen_creditos.append({
            'credito': credito,
            'estado_credito': estado_credito,
            'saldo_actual_credito': saldo_actual_credito
        })

    db.session.commit()

    return render_template(
        'ver_creditos.html',
        resumen_creditos=resumen_creditos,
        sede_actual=sede
    )

@app.route('/ver_creditos_cancelados/<sede>')
def ver_creditos_cancelados(sede):
    if 'user' not in session:
        return redirect('/login')

    sede = sede.strip().upper()

    creditos = Credito.query.filter_by(sede=sede).order_by(Credito.fecha_creacion.desc()).all()
    hoy = date.today()

    resumen_creditos = []

    for credito in creditos:
        actualizar_mora_credito(credito, hoy)

        cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

        if not cuotas:
            continue

        # SOLO mostrar créditos completamente terminados
        if not all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
            continue

        estado_credito = 'CANCELADO'

        resumen_creditos.append({
            'credito': credito,
            'estado_credito': estado_credito
        })

    db.session.commit()

    return render_template(
        'ver_creditos_cancelados.html',
        resumen_creditos=resumen_creditos,
        sede_actual=sede
    )

@app.route('/ver_cuotas/<int:credito_id>')
def ver_cuotas(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    actualizar_mora_credito(credito)
    db.session.commit()

    cuotas = Cuota.query.filter_by(
        credito_id=credito_id
    ).order_by(Cuota.numero).all()

    pagos_por_cuota = {}
    ultimo_pago = None

    for cuota in cuotas:
        pagos = Pago.query.filter(
            Pago.cuota_id == cuota.id,
            Pago.tipo_pago != 'ABONO_CAPITAL'
        ).order_by(Pago.fecha).all()

        pagos_por_cuota[cuota.id] = pagos

        if pagos:
            ultimo_pago_cuota = pagos[-1]
            if ultimo_pago is None or ultimo_pago_cuota.fecha > ultimo_pago.fecha:
                ultimo_pago = ultimo_pago_cuota

    # ABONOS A CAPITAL SEPARADOS DE LOS PAGOS DE CUOTA
    abonos_capital = AbonoCapital.query.filter(
        AbonoCapital.credito_id == credito.id,
        AbonoCapital.activo == True,
        AbonoCapital.reversado == False
    ).order_by(
        AbonoCapital.fecha.asc(),
        AbonoCapital.id.asc()
    ).all()

    abonos_por_cuota = {cuota.id: [] for cuota in cuotas}

    for abono in abonos_capital:
        fecha_abono = abono.fecha.date() if isinstance(abono.fecha, datetime) else abono.fecha

        for cuota in cuotas:
            fecha_cuota = cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago

            if fecha_abono <= fecha_cuota:
                abonos_por_cuota[cuota.id].append(abono)
                break

    hoy = date.today()

    cuotas_exigibles_hoy = []
    for cuota in cuotas:
        fecha_cuota = cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago

        if cuota.estado in ['PENDIENTE', 'EN MORA', 'ABONO'] and fecha_cuota <= hoy:
            cuotas_exigibles_hoy.append(cuota)

    cuota_pendiente_total = round(
        sum((cuota.saldo_pendiente or 0) for cuota in cuotas_exigibles_hoy),
        2
    )

    mora_total = round(
        sum((cuota.interes_mora or 0) for cuota in cuotas_exigibles_hoy),
        2
    )

    deuda_total_fecha = round(cuota_pendiente_total + mora_total, 2)

    if any(c.estado == 'EN MORA' for c in cuotas):
        estado_credito = 'EN MORA'
    elif all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
        estado_credito = 'CANCELADO'
    elif any(c.estado == 'ABONO' for c in cuotas):
        estado_credito = 'CON ABONOS'
    else:
        estado_credito = 'AL DÍA'

    esta_al_dia = deuda_total_fecha <= 0

    historial_tasas = CambioTasaInteresCredito.query.filter_by(
        credito_id=credito.id
    ).order_by(CambioTasaInteresCredito.numero_cuota.asc()).all()

    return render_template(
        'ver_cuotas.html',
        credito=credito,
        cuotas=cuotas,
        pagos_por_cuota=pagos_por_cuota,
        abonos_por_cuota=abonos_por_cuota,
        ultimo_pago=ultimo_pago,
        estado_credito=estado_credito,
        cuota_pendiente_total=cuota_pendiente_total,
        mora_total=mora_total,
        deuda_total_fecha=deuda_total_fecha,
        esta_al_dia=esta_al_dia,
        cuotas_exigibles_hoy=cuotas_exigibles_hoy,
        historial_tasas=historial_tasas
    )

@app.route('/pagar_cuota/<int:cuota_id>', methods=['GET', 'POST'])
def pagar_cuota(cuota_id):
    if 'user' not in session:
        return redirect('/login')

    cuota = Cuota.query.get_or_404(cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)

    if request.method == 'POST':
        fecha_pago = datetime.strptime(request.form['fecha_pago'], '%Y-%m-%d')
        valor_pago = limpiar_valor_moneda(request.form['valor'])
        medio_pago = request.form['medio_pago']
        observacion = request.form.get('observacion', '').strip()

        if medio_pago == 'OTRO':
            medio_pago_otro = request.form.get('medio_pago_otro', '').strip()
            if not medio_pago_otro:
                return "Debes escribir el otro medio de pago"
            medio_pago = medio_pago_otro

        if valor_pago <= 0:
            return "El pago debe ser mayor que cero"

        actualizar_mora_credito(credito, fecha_pago.date())
        db.session.commit()

        cuota = Cuota.query.get_or_404(cuota_id)

        dias_mora_al_pago = cuota.dias_mora or 0
        mora_generada_al_pago = round(cuota.interes_mora or 0, 2)
        saldo_pendiente_antes_pago = round(cuota.saldo_pendiente or 0, 2)

        restante = round(valor_pago, 2)
        hubo_abono_extra_capital = False

        saldo_cuota_hoy = round(cuota.saldo_pendiente or 0, 2)
        mora_hoy = round(cuota.interes_mora or 0, 2)
        total_exigible = round(saldo_cuota_hoy + mora_hoy, 2)

        valor_aplicado_cuota = 0
        valor_aplicado_mora = 0
        valor_aplicado_prepago = 0
        valor_aplicado_interes = 0
        valor_aplicado_capital = 0

        if cuota.saldo_pendiente > 0:
            aplicado_cuota = min(restante, round(cuota.saldo_pendiente, 2))
            valor_aplicado_cuota = round(valor_aplicado_cuota + aplicado_cuota, 2)
            cuota.saldo_pendiente = round(cuota.saldo_pendiente - aplicado_cuota, 2)
            restante = round(restante - aplicado_cuota, 2)

            interes_cuota = round(cuota.interes or 0, 2)

            valor_aplicado_interes = min(valor_aplicado_cuota, interes_cuota)
            valor_aplicado_capital = round(
                max(valor_aplicado_cuota - valor_aplicado_interes, 0),
                2
            )

            if cuota.saldo_pendiente <= 0:
                cuota.saldo_pendiente = 0
                credito.saldo_actual = round(cuota.saldo_restante, 2)

        if restante > 0 and cuota.interes_mora > 0:
            aplicado_mora = min(restante, round(cuota.interes_mora, 2))
            valor_aplicado_mora = round(valor_aplicado_mora + aplicado_mora, 2)
            cuota.interes_mora = round(cuota.interes_mora - aplicado_mora, 2)
            restante = round(restante - aplicado_mora, 2)

        if restante > 0:
            valor_aplicado_prepago = round(restante, 2)
            credito.saldo_actual = round((credito.saldo_actual or 0) - restante, 2)

            if credito.saldo_actual < 0:
                credito.saldo_actual = 0

            if valor_aplicado_prepago >= 1:
                hubo_abono_extra_capital = True

            restante = 0

        cuota.saldo_pendiente = round(max(cuota.saldo_pendiente, 0), 2)
        cuota.interes_mora = round(max(cuota.interes_mora, 0), 2)

        if cuota.saldo_pendiente <= 1:
            cuota.saldo_pendiente = 0

        if cuota.interes_mora <= 1:
            cuota.interes_mora = 0

        if cuota.saldo_pendiente <= 0 and cuota.interes_mora <= 0:
            cuota.saldo_pendiente = 0
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = 0
            cuota.estado = 'PAGADA'

        elif cuota.saldo_pendiente <= 0 and cuota.interes_mora > 0:
            cuota.saldo_pendiente = 0
            cuota.total_cobro = round(cuota.interes_mora, 2)
            cuota.estado = 'EN MORA'

        else:
            cuota.total_cobro = round(cuota.saldo_pendiente + cuota.interes_mora, 2)
            if cuota.dias_mora > 0:
                cuota.estado = 'EN MORA'
            else:
                cuota.estado = 'ABONO'

        pago = Pago(
            cuota_id=cuota.id,
            fecha=fecha_pago,
            valor=valor_pago,
            medio_pago=medio_pago,
            valor_aplicado_interes=round(valor_aplicado_interes, 2),
            valor_aplicado_capital=round(valor_aplicado_capital, 2),
            valor_aplicado_mora=round(valor_aplicado_mora, 2),
            valor_aplicado_prepago_capital=round(valor_aplicado_prepago, 2),
            tipo_pago='PAGO_CUOTA',
            dias_mora_pagados=dias_mora_al_pago,
            mora_generada_al_pago=mora_generada_al_pago,
            saldo_pendiente_antes_pago=saldo_pendiente_antes_pago,
            total_exigible_al_pago=total_exigible,
            observacion=observacion if observacion else "Pago registrado en el sistema financiero"
        )

        db.session.add(pago)
        db.session.flush()

        if hubo_abono_extra_capital:
            if credito.tipo_cuota == 'VARIABLE' and credito.tipo_interes == 'VARIABLE':
                recalcular_cuotas_variables_pendientes(
                    credito=credito,
                    cuota_actual_numero=cuota.numero,
                    fecha_base=cuota.fecha_pago
                )
            else:
                recalcular_cuotas_pendientes(
                    credito=credito,
                    cuota_actual_numero=cuota.numero,
                    fecha_base=cuota.fecha_pago
                )

        db.session.commit()

        return redirect(url_for('ver_recibo_pago', pago_id=pago.id))

    actualizar_mora_credito(credito, datetime.utcnow().date())
    db.session.commit()

    cuota = Cuota.query.get_or_404(cuota_id)

    return render_template('pagar_cuota.html', cuota=cuota)


@app.route('/pagar_deuda_fecha/<int:credito_id>', methods=['GET', 'POST'])
def pagar_deuda_fecha(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if request.method == 'POST':
        fecha_pago = datetime.strptime(request.form['fecha_pago'], '%Y-%m-%d').date()
        valor_pago = limpiar_valor_moneda(request.form['valor'])
        medio_pago = request.form['medio_pago']
        observacion = request.form.get('observacion', '').strip()

        if medio_pago == 'OTRO':
            medio_pago_otro = request.form.get('medio_pago_otro', '').strip()
            if not medio_pago_otro:
                return "Debes escribir el otro medio de pago"
            medio_pago = medio_pago_otro

        if valor_pago <= 0:
            return "El pago debe ser mayor que cero"

        pagos_ids = aplicar_pago_deuda_fecha(
            credito=credito,
            fecha_pago=fecha_pago,
            valor_pago=valor_pago,
            medio_pago=medio_pago
        )

        if not pagos_ids:
            return redirect(f'/ver_cuotas/{credito.id}')

        ids_texto = ",".join(str(x) for x in pagos_ids)
        return redirect(url_for('ver_recibo_deuda_fecha', credito_id=credito.id, pagos=ids_texto))

    actualizar_mora_credito(credito, datetime.utcnow().date())
    db.session.commit()

    cuotas = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    hoy = datetime.utcnow().date()
    cuotas_exigibles_hoy = [
        cuota for cuota in cuotas
        if (cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago) <= hoy
    ]

    cuota_pendiente_total = round(sum((c.saldo_pendiente or 0) for c in cuotas_exigibles_hoy), 2)
    mora_total = round(sum((c.interes_mora or 0) for c in cuotas_exigibles_hoy), 2)
    deuda_total_fecha = round(cuota_pendiente_total + mora_total, 2)

    return render_template(
        'pagar_deuda_fecha.html',
        credito=credito,
        cuotas_exigibles_hoy=cuotas_exigibles_hoy,
        cuota_pendiente_total=cuota_pendiente_total,
        mora_total=mora_total,
        deuda_total_fecha=deuda_total_fecha
    )


@app.route('/abono_capital/<int:credito_id>', methods=['GET', 'POST'])
def abono_capital(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    actualizar_mora_credito(credito, date.today())
    db.session.commit()

    cuotas_credito = Cuota.query.filter_by(
        credito_id=credito.id
    ).order_by(Cuota.numero.asc()).all()


    abonos_capital = AbonoCapital.query.filter_by(
        credito_id=credito.id
    ).order_by(AbonoCapital.fecha.desc()).all()

    cuotas_activas = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    hoy = date.today()
    cuotas_exigibles_hoy = [
        cuota for cuota in cuotas_activas
        if (cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago) <= hoy
    ]

    deuda_total_fecha = round(
        sum((cuota.saldo_pendiente or 0) + (cuota.interes_mora or 0) for cuota in cuotas_exigibles_hoy),
        2
    )

    if request.method == 'POST':
        fecha_pago = datetime.strptime(request.form['fecha_pago'], '%Y-%m-%d')
        fecha_pago_date = fecha_pago.date()

        valor_pago = limpiar_valor_moneda(request.form['valor'])
        medio_pago = request.form['medio_pago']

        if medio_pago == 'OTRO':
            medio_pago_otro = request.form.get('medio_pago_otro', '').strip()
            if not medio_pago_otro:
                return "Debes escribir el otro medio de pago"
            medio_pago = medio_pago_otro

        if valor_pago <= 0:
            return "El abono a capital debe ser mayor que cero"

        saldo_real_credito = round(sum(
            cuota.capital or 0
            for cuota in Cuota.query.filter(
                Cuota.credito_id == credito.id,
                Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
            ).all()
        ), 2)

        if credito.saldo_actual is None or credito.saldo_actual <= 0:
            credito.saldo_actual = saldo_real_credito
            db.session.flush()

        if valor_pago > credito.saldo_actual:
            from flask import flash

            flash("El abono a capital no puede ser mayor al saldo actual del crédito", "abono_error")
            return redirect(url_for('abono_capital', credito_id=credito.id))

        # Validar deuda según la fecha real del abono
        actualizar_mora_credito(credito, fecha_pago_date)
        db.session.flush()

        cuotas_en_fecha = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
        ).order_by(Cuota.numero).all()

        cuotas_exigibles_fecha = [
            cuota for cuota in cuotas_en_fecha
            if (cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago) <= fecha_pago_date
        ]

        deuda_en_fecha_abono = round(
            sum((cuota.saldo_pendiente or 0) + (cuota.interes_mora or 0) for cuota in cuotas_exigibles_fecha),
            2
        )

        if deuda_en_fecha_abono > 0:
            db.session.rollback()
            return f"No se puede hacer abono a capital porque en la fecha seleccionada existía deuda exigible por {formato_cop(deuda_en_fecha_abono)}"

        saldo_antes_abono = round(credito.saldo_actual or 0, 2)

        credito.saldo_actual = round(credito.saldo_actual - valor_pago, 2)

        if credito.saldo_actual < 0:
            credito.saldo_actual = 0
        cuota_referencia = None

        for cuota in cuotas_en_fecha:
            fecha_cuota = cuota.fecha_pago.date() if isinstance(cuota.fecha_pago, datetime) else cuota.fecha_pago

            if fecha_cuota > fecha_pago_date and cuota.estado in ['PENDIENTE', 'EN MORA', 'ABONO']:
                cuota_referencia = cuota
                break

        if not cuota_referencia:
            db.session.rollback()
            return "No se encontró una cuota futura para aplicar el abono a capital"

        abono = AbonoCapital(
            credito_id=credito.id,
            fecha=fecha_pago,
            valor=valor_pago,
            medio_pago=medio_pago,
            observacion="ABONO A CAPITAL"
        )
        db.session.add(abono)
        db.session.flush()

        if credito.tipo_cuota == 'VARIABLE' and credito.tipo_interes == 'VARIABLE':
            recalcular_cuotas_variables_pendientes(
                credito=credito,
                cuota_actual_numero=cuota_referencia.numero - 1,
                fecha_base=sumar_meses(cuota_referencia.fecha_pago, -1)
            )
        else:
            recalcular_cuotas_pendientes(
                credito=credito,
                cuota_actual_numero=cuota_referencia.numero - 1,
                fecha_base=sumar_meses(cuota_referencia.fecha_pago, -1)
            )

        actualizar_mora_credito(credito, date.today())

        db.session.commit()

        return redirect(url_for("ver_recibo_abono_capital", abono_id=abono.id))

    return render_template(
        'abono_capital.html',
        credito=credito,
        deuda_total_fecha=deuda_total_fecha,
        abonos_capital=abonos_capital
    )

@app.route('/configuracion_tasa', methods=['GET', 'POST'])
def configuracion_tasa():
    if 'user' not in session:
        return redirect('/login')

    if request.method == 'POST':
        anio = int(request.form['anio'])
        mes = int(request.form['mes'])
        tasa_anual = float(str(request.form['tasa_anual']).replace(',', '.'))

        tasa_mensual = convertir_tasa_anual_a_mensual(tasa_anual)
        tasa_diaria = convertir_tasa_mensual_a_diaria(tasa_mensual)

        tasa_periodo = TasaPeriodo.query.filter_by(anio=anio, mes=mes).first()

        if not tasa_periodo:
            tasa_periodo = TasaPeriodo(
                anio=anio,
                mes=mes,
                tasa_anual=tasa_anual,
                tasa_mensual=tasa_mensual,
                tasa_diaria=tasa_diaria
            )
            db.session.add(tasa_periodo)
        else:
            tasa_periodo.tasa_anual = tasa_anual
            tasa_periodo.tasa_mensual = tasa_mensual
            tasa_periodo.tasa_diaria = tasa_diaria

        # respaldo global opcional
        configuracion = ConfiguracionTasa.query.first()
        if not configuracion:
            configuracion = ConfiguracionTasa(
                nombre='TASA_MORA',
                tasa_anual=tasa_anual,
                tasa_mensual=tasa_mensual,
                tasa_diaria=tasa_diaria
            )
            db.session.add(configuracion)
        else:
            configuracion.tasa_anual = tasa_anual
            configuracion.tasa_mensual = tasa_mensual
            configuracion.tasa_diaria = tasa_diaria

        db.session.commit()
        flash(f'Tasa guardada correctamente para {mes}/{anio}', 'success')
        return redirect(url_for('configuracion_tasa'))

    tasas = TasaPeriodo.query.order_by(TasaPeriodo.anio.desc(), TasaPeriodo.mes.asc()).all()

    return render_template(
        'configuracion_tasa.html',
        tasas=tasas
    )


@app.route('/liquidar_credito/<int:credito_id>', methods=['GET', 'POST'])
def liquidar_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    cuotas_activas = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    if not cuotas_activas:
        return "Este crédito ya se encuentra liquidado"

    if request.method == 'POST':
        fecha_pago = datetime.strptime(request.form['fecha_pago'], '%Y-%m-%d')
        valor_pago = limpiar_valor_moneda(request.form['valor'])
        medio_pago = request.form['medio_pago']

        if medio_pago == 'OTRO':
            medio_pago_otro = request.form.get('medio_pago_otro', '').strip()
            if not medio_pago_otro:
                return "Debes escribir el otro medio de pago"
            medio_pago = medio_pago_otro

        if valor_pago <= 0:
            return "El valor de la liquidación debe ser mayor que cero"

        actualizar_mora_credito(credito, fecha_pago.date())
        db.session.commit()

        cuotas_activas = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
        ).order_by(Cuota.numero).all()

        if not cuotas_activas:
            return "Este crédito ya se encuentra liquidado"

        componentes = calcular_componentes_liquidacion(credito, fecha_pago.date())

        cuota_actual = componentes['cuota_actual']
        capital_insoluto = componentes['capital_insoluto']
        interes_corriente = componentes['interes_corriente']
        total_mora = componentes['total_mora']
        total_liquidacion = componentes['total_liquidacion']

        if round(valor_pago) != round(total_liquidacion):
            return (
                f"El valor ingresado no coincide con la liquidación exacta. "
                f"Debes pagar {round(total_liquidacion)}"
            )

        pago = Pago(
            cuota_id=cuota_actual.id,
            fecha=fecha_pago,
            valor=valor_pago,
            medio_pago=medio_pago,
            valor_aplicado_mora=round(total_mora, 2),
            valor_aplicado_interes=round(interes_corriente, 2),
            valor_aplicado_capital=round(capital_insoluto, 2),
            valor_aplicado_prepago_capital=0,
            tipo_pago='LIQUIDACION_TOTAL',
            dias_mora_pagados=cuota_actual.dias_mora or 0,
            mora_generada_al_pago=total_mora,
            saldo_pendiente_antes_pago=capital_insoluto,
            total_exigible_al_pago=total_liquidacion,
            observacion='LIQUIDACION TOTAL DEL CREDITO'
        )
        db.session.add(pago)

        for cuota in cuotas_activas:
            cuota.saldo_pendiente = 0
            cuota.dias_mora = 0
            cuota.interes_mora = 0
            cuota.total_cobro = 0
            cuota.estado = 'LIQUIDADA'

        credito.saldo_actual = 0
        credito.cuota_mensual = 0

        if hasattr(credito, 'fecha_liquidacion'):
            credito.fecha_liquidacion = fecha_pago
        if hasattr(credito, 'valor_liquidado'):
            credito.valor_liquidado = valor_pago

        db.session.commit()
        return redirect(url_for('ver_recibo_pago', pago_id=pago.id))

    actualizar_mora_credito(credito, datetime.utcnow().date())
    db.session.commit()

    cuotas_activas = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado.in_(['PENDIENTE', 'EN MORA', 'ABONO'])
    ).order_by(Cuota.numero).all()

    if not cuotas_activas:
        return "Este crédito ya se encuentra liquidado"

    componentes = calcular_componentes_liquidacion(credito, datetime.utcnow().date())

    cuota_actual = componentes['cuota_actual']
    capital_insoluto = componentes['capital_insoluto']
    interes_corriente = componentes['interes_corriente']
    total_mora = componentes['total_mora']
    total_liquidacion = componentes['total_liquidacion']

    return render_template(
        'liquidar_credito.html',
        credito=credito,
        cuota_actual=cuota_actual,
        capital_insoluto=capital_insoluto,
        interes_corriente=interes_corriente,
        total_mora=total_mora,
        total_liquidacion=total_liquidacion
    )

def construir_datos_reporte(anio_seleccionado, sede_seleccionada):
    def fecha_solo_fecha(valor):
        if valor is None:
            return None
        if isinstance(valor, datetime):
            return valor.date()
        return valor

    def sede_normalizada(sede):
        if not sede:
            return 'CRV'
        sede = str(sede).strip().upper()
        if sede == 'SAS':
            return 'CRV'
        return sede

    sedes_db = db.session.query(Credito.sede).distinct().all()
    sedes_base = sorted(list(set([
        sede_normalizada(s[0]) for s in sedes_db if s[0]
    ])))

    if not sedes_base:
        sedes_base = ['CRV']

    if sede_seleccionada == 'TODAS':
        sedes_filtradas = sedes_base
    else:
        sedes_filtradas = [sede_seleccionada]

    def credito_aplica(credito):
        return sede_normalizada(credito.sede) in sedes_filtradas

    def total_inyecciones_credito(credito_id, anio=None, mes=None):
        query = db.session.query(
            db.func.coalesce(db.func.sum(InyeccionCapital.valor), 0)
        ).filter(
            InyeccionCapital.credito_id == credito_id
        )

        if anio:
            query = query.filter(db.extract('year', InyeccionCapital.fecha) == anio)

        if mes:
            query = query.filter(db.extract('month', InyeccionCapital.fecha) == mes)

        return round(query.scalar() or 0, 2)

    def total_abonos_credito(credito_id, anio=None, mes=None):
        query = db.session.query(
            db.func.coalesce(db.func.sum(AbonoCapital.valor), 0)
        ).filter(
            AbonoCapital.credito_id == credito_id,
            AbonoCapital.activo == True
        )

        if anio:
            query = query.filter(db.extract('year', AbonoCapital.fecha) == anio)

        if mes:
            query = query.filter(db.extract('month', AbonoCapital.fecha) == mes)

        return round(query.scalar() or 0, 2)

    creditos = Credito.query.all()
    creditos_filtrados = [c for c in creditos if credito_aplica(c)]

    creditos_anio = [
        c for c in creditos_filtrados
        if fecha_solo_fecha(c.fecha_creacion)
        and fecha_solo_fecha(c.fecha_creacion).year == anio_seleccionado
    ]

    pagos_anio = (
        db.session.query(Pago, Cuota, Credito)
        .join(Cuota, Pago.cuota_id == Cuota.id)
        .join(Credito, Cuota.credito_id == Credito.id)
        .filter(
            Pago.activo == True,
            db.extract('year', Pago.fecha) == anio_seleccionado
        )
        .all()
    )

    pagos_anio = [
        (p, q, c) for p, q, c in pagos_anio
        if credito_aplica(c)
    ]

    cuotas_anio = (
        db.session.query(Cuota, Credito)
        .join(Credito, Cuota.credito_id == Credito.id)
        .filter(
            db.extract('year', Cuota.fecha_pago) == anio_seleccionado
        )
        .all()
    )

    cuotas_anio = [
        (q, c) for q, c in cuotas_anio
        if credito_aplica(c)
    ]

    # ===============================
    # RESUMEN GENERAL
    # ===============================
    total_prestado_creditos = round(sum(c.monto_financiado or 0 for c in creditos_anio), 2)

    total_inyecciones_anio = round(sum(
        total_inyecciones_credito(c.id, anio_seleccionado)
        for c in creditos_filtrados
    ), 2)

    total_prestado = round(total_prestado_creditos + total_inyecciones_anio, 2)

    total_pagos_anio = round(sum(p.valor or 0 for p, q, c in pagos_anio), 2)

    total_abonos_anio = round(sum(
        total_abonos_credito(c.id, anio_seleccionado)
        for c in creditos_filtrados
    ), 2)

    total_recaudado = round(total_pagos_anio + total_abonos_anio, 2)

    saldo_actual_total = round(sum(c.saldo_actual or 0 for c in creditos_filtrados), 2)

    interes_corriente_causado = round(sum(q.interes or 0 for q, c in cuotas_anio), 2)
    mora_causada = round(sum(q.interes_mora or 0 for q, c in cuotas_anio), 2)

    interes_corriente_recaudado = round(sum(
        p.valor_aplicado_interes or 0 for p, q, c in pagos_anio
    ), 2)

    mora_recaudada = round(sum(
        p.valor_aplicado_mora or 0 for p, q, c in pagos_anio
    ), 2)

    diferencia_interes_corriente = round(
        interes_corriente_causado - interes_corriente_recaudado,
        2
    )

    diferencia_mora = round(
        mora_causada - mora_recaudada,
        2
    )

    diferencia_total = round(diferencia_interes_corriente + diferencia_mora, 2)

    resumen_general = {
        'total_prestado': total_prestado,
        'total_recaudado': total_recaudado,
        'saldo_actual_total': saldo_actual_total,
        'interes_corriente_causado': interes_corriente_causado,
        'interes_corriente_recaudado': interes_corriente_recaudado,
        'mora_causada': mora_causada,
        'mora_recaudada': mora_recaudada,
        'diferencia_interes_corriente': diferencia_interes_corriente,
        'diferencia_mora': diferencia_mora,
        'diferencia_total': diferencia_total
    }

    # ===============================
    # RESUMEN POR SEDE
    # ===============================
    resumen_por_sede = []

    for sede in sedes_filtradas:
        creditos_sede = [
            c for c in creditos_filtrados
            if sede_normalizada(c.sede) == sede
        ]

        creditos_sede_anio = [
            c for c in creditos_sede
            if fecha_solo_fecha(c.fecha_creacion)
            and fecha_solo_fecha(c.fecha_creacion).year == anio_seleccionado
        ]

        pagos_sede_anio = [
            (p, q, c) for p, q, c in pagos_anio
            if sede_normalizada(c.sede) == sede
        ]

        cuotas_sede_anio = [
            (q, c) for q, c in cuotas_anio
            if sede_normalizada(c.sede) == sede
        ]

        prestado_sede = round(
            sum(c.monto_financiado or 0 for c in creditos_sede_anio)
            + sum(total_inyecciones_credito(c.id, anio_seleccionado) for c in creditos_sede),
            2
        )

        recaudado_sede = round(
            sum(p.valor or 0 for p, q, c in pagos_sede_anio)
            + sum(total_abonos_credito(c.id, anio_seleccionado) for c in creditos_sede),
            2
        )

        interes_causado_sede = round(sum(q.interes or 0 for q, c in cuotas_sede_anio), 2)
        mora_causada_sede = round(sum(q.interes_mora or 0 for q, c in cuotas_sede_anio), 2)

        interes_recaudado_sede = round(sum(
            p.valor_aplicado_interes or 0 for p, q, c in pagos_sede_anio
        ), 2)

        mora_recaudada_sede = round(sum(
            p.valor_aplicado_mora or 0 for p, q, c in pagos_sede_anio
        ), 2)

        diferencia_interes_sede = round(interes_causado_sede - interes_recaudado_sede, 2)
        diferencia_mora_sede = round(mora_causada_sede - mora_recaudada_sede, 2)

        resumen_por_sede.append({
            'sede': sede,
            'total_prestado': prestado_sede,
            'total_recaudado': recaudado_sede,
            'saldo_actual': round(sum(c.saldo_actual or 0 for c in creditos_sede), 2),
            'interes_corriente_causado': interes_causado_sede,
            'interes_corriente_recaudado': interes_recaudado_sede,
            'mora_causada': mora_causada_sede,
            'mora_recaudada': mora_recaudada_sede,
            'diferencia_interes_corriente': diferencia_interes_sede,
            'diferencia_mora': diferencia_mora_sede,
            'diferencia_total': round(diferencia_interes_sede + diferencia_mora_sede, 2)
        })

    # ===============================
    # RESUMEN MENSUAL
    # ===============================
    resumen_mensual = []

    for mes in range(1, 13):
        pagos_mes = [
            (p, q, c) for p, q, c in pagos_anio
            if fecha_solo_fecha(p.fecha)
            and fecha_solo_fecha(p.fecha).month == mes
        ]

        cuotas_mes = [
            (q, c) for q, c in cuotas_anio
            if fecha_solo_fecha(q.fecha_pago)
            and fecha_solo_fecha(q.fecha_pago).month == mes
        ]

        abonos_mes = round(sum(
            total_abonos_credito(c.id, anio_seleccionado, mes)
            for c in creditos_filtrados
        ), 2)

        interes_causado_mes = round(sum(q.interes or 0 for q, c in cuotas_mes), 2)
        mora_causada_mes = round(sum(q.interes_mora or 0 for q, c in cuotas_mes), 2)

        interes_recaudado_mes = round(sum(
            p.valor_aplicado_interes or 0 for p, q, c in pagos_mes
        ), 2)

        mora_recaudada_mes = round(sum(
            p.valor_aplicado_mora or 0 for p, q, c in pagos_mes
        ), 2)

        total_ingresos_mes = round(
            sum(p.valor or 0 for p, q, c in pagos_mes) + abonos_mes,
            2
        )

        resumen_mensual.append({
            'mes': MESES_ES[mes],
            'interes_corriente_causado': interes_causado_mes,
            'interes_corriente_recaudado': interes_recaudado_mes,
            'mora_causada': mora_causada_mes,
            'mora_recaudada': mora_recaudada_mes,
            'diferencia_interes_corriente': round(interes_causado_mes - interes_recaudado_mes, 2),
            'diferencia_mora': round(mora_causada_mes - mora_recaudada_mes, 2),
            'total_ingresos': total_ingresos_mes
        })

    # ===============================
    # TABLAS DETALLADAS
    # Mantiene IBAGUE/GIRARDOT/ESPINAL/CRV para no romper el HTML actual.
    # ===============================
    sedes_tabla = ['IBAGUE', 'GIRARDOT', 'ESPINAL', 'CRV']

    def crear_tabla_por_mes(campo):
        tabla = []

        for mes in range(1, 13):
            fila = {'mes': MESES_ES[mes]}

            total = 0

            for sede in sedes_tabla:
                valor = 0

                if campo == 'interes_causado':
                    valor = sum(
                        q.interes or 0
                        for q, c in cuotas_anio
                        if sede_normalizada(c.sede) == sede
                        and fecha_solo_fecha(q.fecha_pago)
                        and fecha_solo_fecha(q.fecha_pago).month == mes
                    )

                elif campo == 'interes_recaudado':
                    valor = sum(
                        p.valor_aplicado_interes or 0
                        for p, q, c in pagos_anio
                        if sede_normalizada(c.sede) == sede
                        and fecha_solo_fecha(p.fecha)
                        and fecha_solo_fecha(p.fecha).month == mes
                    )

                elif campo == 'mora_causada':
                    valor = sum(
                        q.interes_mora or 0
                        for q, c in cuotas_anio
                        if sede_normalizada(c.sede) == sede
                        and fecha_solo_fecha(q.fecha_pago)
                        and fecha_solo_fecha(q.fecha_pago).month == mes
                    )

                elif campo == 'mora_recaudada':
                    valor = sum(
                        p.valor_aplicado_mora or 0
                        for p, q, c in pagos_anio
                        if sede_normalizada(c.sede) == sede
                        and fecha_solo_fecha(p.fecha)
                        and fecha_solo_fecha(p.fecha).month == mes
                    )

                valor = round(valor, 2)
                fila[sede] = valor
                total += valor

            fila['TOTAL'] = round(total, 2)
            tabla.append(fila)

        return tabla

    tabla_intereses_causados = crear_tabla_por_mes('interes_causado')
    tabla_intereses_recaudados = crear_tabla_por_mes('interes_recaudado')
    tabla_mora_causada = crear_tabla_por_mes('mora_causada')
    tabla_mora_recaudada = crear_tabla_por_mes('mora_recaudada')

    tabla_diferencia_intereses = []
    tabla_diferencia_mora = []

    for i in range(12):
        fila_int = {'mes': MESES_ES[i + 1]}
        fila_mora = {'mes': MESES_ES[i + 1]}

        total_int = 0
        total_mora = 0

        for sede in sedes_tabla:
            dif_int = round(
                tabla_intereses_causados[i][sede] - tabla_intereses_recaudados[i][sede],
                2
            )

            dif_mora = round(
                tabla_mora_causada[i][sede] - tabla_mora_recaudada[i][sede],
                2
            )

            fila_int[sede] = dif_int
            fila_mora[sede] = dif_mora
            total_int += dif_int
            total_mora += dif_mora

        fila_int['TOTAL'] = round(total_int, 2)
        fila_mora['TOTAL'] = round(total_mora, 2)

        tabla_diferencia_intereses.append(fila_int)
        tabla_diferencia_mora.append(fila_mora)

    def totales_tabla(filas):
        return {
            'IBAGUE': round(sum(f['IBAGUE'] for f in filas), 2),
            'GIRARDOT': round(sum(f['GIRARDOT'] for f in filas), 2),
            'ESPINAL': round(sum(f['ESPINAL'] for f in filas), 2),
            'CRV': round(sum(f['CRV'] for f in filas), 2),
            'TOTAL': round(sum(f['TOTAL'] for f in filas), 2),
        }

    labels_sedes = [fila['sede'] for fila in resumen_por_sede]
    saldo_actual_sedes = [fila['saldo_actual'] for fila in resumen_por_sede]
    total_prestado_sedes = [fila['total_prestado'] for fila in resumen_por_sede]
    total_recaudado_sedes = [fila['total_recaudado'] for fila in resumen_por_sede]
    interes_causado_sedes = [fila['interes_corriente_causado'] for fila in resumen_por_sede]
    interes_recaudado_sedes = [fila['interes_corriente_recaudado'] for fila in resumen_por_sede]
    mora_causada_sedes = [fila['mora_causada'] for fila in resumen_por_sede]
    mora_recaudada_sedes = [fila['mora_recaudada'] for fila in resumen_por_sede]
    diferencia_interes_sedes = [fila['diferencia_interes_corriente'] for fila in resumen_por_sede]
    diferencia_mora_sedes = [fila['diferencia_mora'] for fila in resumen_por_sede]

    labels_meses = [fila['mes'] for fila in resumen_mensual]
    interes_causado_meses = [fila['interes_corriente_causado'] for fila in resumen_mensual]
    interes_recaudado_meses = [fila['interes_corriente_recaudado'] for fila in resumen_mensual]
    mora_causada_meses = [fila['mora_causada'] for fila in resumen_mensual]
    mora_recaudada_meses = [fila['mora_recaudada'] for fila in resumen_mensual]
    diferencia_interes_meses = [fila['diferencia_interes_corriente'] for fila in resumen_mensual]
    diferencia_mora_meses = [fila['diferencia_mora'] for fila in resumen_mensual]
    total_ingresos_meses = [fila['total_ingresos'] for fila in resumen_mensual]

    return {
        'resumen_general': resumen_general,
        'resumen_por_sede': resumen_por_sede,
        'resumen_mensual': resumen_mensual,

        'tabla_intereses_causados': tabla_intereses_causados,
        'tabla_intereses_recaudados': tabla_intereses_recaudados,
        'tabla_mora_causada': tabla_mora_causada,
        'tabla_mora_recaudada': tabla_mora_recaudada,
        'tabla_diferencia_intereses': tabla_diferencia_intereses,
        'tabla_diferencia_mora': tabla_diferencia_mora,

        'totales_intereses_causados': totales_tabla(tabla_intereses_causados),
        'totales_intereses_recaudados': totales_tabla(tabla_intereses_recaudados),
        'totales_mora_causada': totales_tabla(tabla_mora_causada),
        'totales_mora_recaudada': totales_tabla(tabla_mora_recaudada),
        'totales_diferencia_intereses': totales_tabla(tabla_diferencia_intereses),
        'totales_diferencia_mora': totales_tabla(tabla_diferencia_mora),

        'labels_sedes': labels_sedes,
        'saldo_actual_sedes': saldo_actual_sedes,
        'total_prestado_sedes': total_prestado_sedes,
        'total_recaudado_sedes': total_recaudado_sedes,
        'interes_causado_sedes': interes_causado_sedes,
        'interes_recaudado_sedes': interes_recaudado_sedes,
        'mora_causada_sedes': mora_causada_sedes,
        'mora_recaudada_sedes': mora_recaudada_sedes,
        'diferencia_interes_sedes': diferencia_interes_sedes,
        'diferencia_mora_sedes': diferencia_mora_sedes,

        'labels_meses': labels_meses,
        'interes_causado_meses': interes_causado_meses,
        'interes_recaudado_meses': interes_recaudado_meses,
        'mora_causada_meses': mora_causada_meses,
        'mora_recaudada_meses': mora_recaudada_meses,
        'diferencia_interes_meses': diferencia_interes_meses,
        'diferencia_mora_meses': diferencia_mora_meses,
        'total_ingresos_meses': total_ingresos_meses
    }

@app.route('/reporte_financiero')
def reporte_financiero():
    if 'user' not in session:
        return redirect('/login')

    anio_actual = date.today().year
    anio_seleccionado = request.args.get('anio', type=int) or anio_actual
    sede_seleccionada = request.args.get('sede', default='TODAS', type=str).strip().upper()

    datos = construir_datos_reporte(anio_seleccionado, sede_seleccionada)

    # =========================
    # SEDES DINÁMICAS
    # =========================
    sedes_db = db.session.query(Credito.sede).distinct().all()

    sedes_disponibles = ['TODAS'] + sorted([
        s[0] for s in sedes_db if s[0]
    ])

    # =========================
    # REPORTE ACUMULADO GENERAL
    # Incluye:
    # - créditos
    # - inyecciones de capital
    # - pagos normales
    # - abonos a capital activos
    # =========================
    creditos_acumulados = Credito.query.all()

    acumulado_por_sede = {}

    for credito in creditos_acumulados:
        sede = credito.sede or "SIN SEDE"

        if sede not in acumulado_por_sede:
            acumulado_por_sede[sede] = {
                "sede": sede,
                "total_prestamo": 0,
                "total_pagado": 0,
                "total_deben": 0
            }

        total_inyecciones = db.session.query(
            db.func.coalesce(db.func.sum(InyeccionCapital.valor), 0)
        ).filter(
            InyeccionCapital.credito_id == credito.id
        ).scalar() or 0

        total_pagos = db.session.query(
            db.func.coalesce(db.func.sum(Pago.valor), 0)
        ).join(
            Cuota, Pago.cuota_id == Cuota.id
        ).filter(
            Cuota.credito_id == credito.id,
            Pago.activo == True,
            Pago.reversado == False
        ).scalar() or 0

        total_abonos_capital = db.session.query(
            db.func.coalesce(db.func.sum(AbonoCapital.valor), 0)
        ).filter(
            AbonoCapital.credito_id == credito.id,
            AbonoCapital.activo == True,
            AbonoCapital.reversado == False
        ).scalar() or 0

        total_prestamo = round((credito.monto_financiado or 0) + total_inyecciones, 2)
        total_pagado = round(total_pagos + total_abonos_capital, 2)
        total_deben = round(total_prestamo - total_pagado, 2)

        if total_deben < 0:
            total_deben = 0

        acumulado_por_sede[sede]["total_prestamo"] += total_prestamo
        acumulado_por_sede[sede]["total_pagado"] += total_pagado
        acumulado_por_sede[sede]["total_deben"] += total_deben

    reporte_acumulado = list(acumulado_por_sede.values())

    total_prestamo_acumulado = sum(f["total_prestamo"] for f in reporte_acumulado)
    total_pagado_acumulado = sum(f["total_pagado"] for f in reporte_acumulado)
    total_deben_acumulado = sum(f["total_deben"] for f in reporte_acumulado)

    return render_template(
        'reporte_financiero.html',
        anio_actual=anio_actual,
        anio_seleccionado=anio_seleccionado,
        anios_disponibles=list(range(2022, anio_actual + 2)),
        sede_seleccionada=sede_seleccionada,
        sedes_disponibles=sedes_disponibles,
        reporte_acumulado=reporte_acumulado,
        total_prestamo_acumulado=total_prestamo_acumulado,
        total_pagado_acumulado=total_pagado_acumulado,
        total_deben_acumulado=total_deben_acumulado,
        **datos
    )

@app.route('/reporte_financiero/excel')
def exportar_reporte_excel():
    if 'user' not in session:
        return redirect('/login')

    anio_actual = date.today().year
    anio_seleccionado = request.args.get('anio', type=int) or anio_actual
    sede_seleccionada = request.args.get('sede', default='TODAS', type=str).strip().upper()

    datos = construir_datos_reporte(anio_seleccionado, sede_seleccionada)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen General"

    # Estilos básicos
    fill_header = PatternFill("solid", fgColor="1F4E78")
    fill_sub = PatternFill("solid", fgColor="D9EAF7")
    font_white = Font(color="FFFFFF", bold=True)
    font_bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Título
    ws1["A1"] = f"Reporte Financiero - Año {anio_seleccionado} - Sede {sede_seleccionada}"
    ws1["A1"].font = Font(bold=True, size=14)

    # Resumen general
    ws1["A3"] = "Concepto"
    ws1["B3"] = "Valor"
    for c in ["A3", "B3"]:
        ws1[c].fill = fill_header
        ws1[c].font = font_white
        ws1[c].alignment = center

    resumen_general = datos["resumen_general"]
    filas_general = [
        ("Total prestado", resumen_general["total_prestado"]),
        ("Total recaudado", resumen_general["total_recaudado"]),
        ("Saldo actual total", resumen_general["saldo_actual_total"]),
        ("Interés corriente causado", resumen_general["interes_corriente_causado"]),
        ("Interés corriente recaudado", resumen_general["interes_corriente_recaudado"]),
        ("Mora causada", resumen_general["mora_causada"]),
        ("Mora recaudada", resumen_general["mora_recaudada"]),
    ]

    fila = 4
    for concepto, valor in filas_general:
        ws1[f"A{fila}"] = concepto
        ws1[f"B{fila}"] = valor
        ws1[f"B{fila}"].number_format = '$ #,##0'
        fila += 1

    # Hoja por sede
    ws2 = wb.create_sheet("Por Sede")
    headers_sede = [
        "Sede", "Total prestado", "Total recaudado", "Saldo actual",
        "Interés corriente causado", "Interés corriente recaudado",
        "Mora causada", "Mora recaudada"
    ]
    ws2.append(headers_sede)
    for col in range(1, len(headers_sede) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_white
        cell.alignment = center

    for item in datos["resumen_por_sede"]:
        ws2.append([
            item["sede"],
            item["total_prestado"],
            item["total_recaudado"],
            item["saldo_actual"],
            item["interes_corriente_causado"],
            item["interes_corriente_recaudado"],
            item["mora_causada"],
            item["mora_recaudada"],
        ])

    for row in ws2.iter_rows(min_row=2, min_col=2, max_col=8):
        for cell in row:
            cell.number_format = '$ #,##0'

    # Hoja mensual
    ws3 = wb.create_sheet("Resumen Mensual")
    headers_mes = [
        "Mes", "Interés corriente causado", "Interés corriente recaudado",
        "Mora causada", "Mora recaudada", "Total ingresos"
    ]
    ws3.append(headers_mes)
    for col in range(1, len(headers_mes) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_white
        cell.alignment = center

    for item in datos["resumen_mensual"]:
        ws3.append([
            item["mes"],
            item["interes_corriente_causado"],
            item["interes_corriente_recaudado"],
            item["mora_causada"],
            item["mora_recaudada"],
            item["total_ingresos"],
        ])

    for row in ws3.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = '$ #,##0'

    # Anchos
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, 28)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre = f"reporte_financiero_{anio_seleccionado}_{sede_seleccionada}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/reporte_financiero/pdf')
def exportar_reporte_pdf():
    if 'user' not in session:
        return redirect('/login')

    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import os

    anio_actual = date.today().year
    anio_seleccionado = request.args.get('anio', type=int) or anio_actual
    sede_seleccionada = request.args.get('sede', default='TODAS', type=str).strip().upper()

    datos = construir_datos_reporte(anio_seleccionado, sede_seleccionada)

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    azul = colors.HexColor("#0b2f4f")
    dorado = colors.HexColor("#d6a21e")
    gris_claro = colors.HexColor("#f4f7fb")
    borde = colors.HexColor("#d9e2ec")
    verde = colors.HexColor("#15803d")
    rojo = colors.HexColor("#b91c1c")
    naranja = colors.HexColor("#c76a00")

    titulo_style = ParagraphStyle(
        "TituloCRV",
        parent=styles["Title"],
        fontSize=22,
        textColor=azul,
        alignment=TA_CENTER,
        spaceAfter=8
    )

    subtitulo_style = ParagraphStyle(
        "SubtituloCRV",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#334155"),
        alignment=TA_CENTER,
        spaceAfter=12
    )

    seccion_style = ParagraphStyle(
        "SeccionCRV",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=azul,
        spaceBefore=12,
        spaceAfter=8
    )

    normal_style = ParagraphStyle(
        "NormalCRV",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#1f2937"),
        alignment=TA_LEFT
    )

    elementos = []

    def agregar_encabezado():
        logo_path = os.path.join(app.static_folder, "logo.png")

        logo = ""
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=95, height=58)

        empresa = Paragraph("""
            <b>CONSTRUCCIONES Y URBANIZACIONES S.A.S</b><br/>
            NIT: 901.527.083-2<br/>
            AV. AMBALÁ N° 27-136 - PISO 3<br/>
            IBAGUÉ - TOLIMA<br/>
            TELÉFONO: 311 414 5843
        """, normal_style)

        titulo = Paragraph("REPORTE FINANCIERO", titulo_style)
        subtitulo = Paragraph(
            f"Consolidado financiero - Año {anio_seleccionado} - Sede {sede_seleccionada}",
            subtitulo_style
        )

        tabla_header = Table(
            [[logo, [titulo, subtitulo], empresa]],
            colWidths=[140, 360, 230]
        )

        tabla_header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("LINEBELOW", (0, 0), (-1, -1), 1, borde),
        ]))

        elementos.append(tabla_header)
        elementos.append(Spacer(1, 14))

    def card(titulo, valor, color_fondo):
        return [
            Paragraph(f"<font color='white'><b>{titulo}</b></font>", normal_style),
            Paragraph(f"<font color='white' size='14'><b>{formato_cop(valor)}</b></font>", normal_style)
        ]

    def tabla_resumen_general():
        r = datos["resumen_general"]

        data = [
            [
                card("TOTAL PRESTADO", r["total_prestado"], azul),
                card("TOTAL RECAUDADO", r["total_recaudado"], verde),
                card("SALDO ACTUAL TOTAL", r["saldo_actual_total"], naranja),
                card("INTERÉS CAUSADO", r["interes_corriente_causado"], colors.HexColor("#6d28d9")),
            ],
            [
                card("INTERÉS RECAUDADO", r["interes_corriente_recaudado"], colors.HexColor("#0e7490")),
                card("MORA CAUSADA", r["mora_causada"], rojo),
                card("MORA RECAUDADA", r["mora_recaudada"], colors.HexColor("#1e293b")),
                card("DIFERENCIA TOTAL", r["diferencia_total"], azul),
            ]
        ]

        t = Table(data, colWidths=[180, 180, 180, 180], rowHeights=[65, 65])

        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#1d4ed8")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#15803d")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#d97706")),
            ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#6d28d9")),
            ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#0e7490")),
            ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#b91c1c")),
            ("BACKGROUND", (2, 1), (2, 1), colors.HexColor("#1e293b")),
            ("BACKGROUND", (3, 1), (3, 1), colors.HexColor("#0f172a")),
            ("BOX", (0, 0), (-1, -1), 0.5, borde),
            ("INNERGRID", (0, 0), (-1, -1), 8, colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ]))

        elementos.append(Paragraph("Resumen general anual", seccion_style))
        elementos.append(t)
        elementos.append(Spacer(1, 14))

    def tabla_normal(titulo, headers, filas, col_widths=None):
        elementos.append(Paragraph(titulo, seccion_style))

        data = [headers] + filas

        if col_widths is None:
            col_widths = [90] * len(headers)

        t = Table(data, repeatRows=1, colWidths=col_widths)

        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), azul),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, borde),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, gris_claro]),
        ]

        t.setStyle(TableStyle(estilo))
        elementos.append(t)
        elementos.append(Spacer(1, 12))

    agregar_encabezado()
    tabla_resumen_general()

    filas_sede = []
    for item in datos["resumen_por_sede"]:
        filas_sede.append([
            item["sede"],
            formato_cop(item["total_prestado"]),
            formato_cop(item["total_recaudado"]),
            formato_cop(item["saldo_actual"]),
            formato_cop(item["interes_corriente_causado"]),
            formato_cop(item["interes_corriente_recaudado"]),
            formato_cop(item["mora_causada"]),
            formato_cop(item["mora_recaudada"]),
            formato_cop(item["diferencia_total"]),
        ])

    tabla_normal(
        "Resumen por sede",
        [
            "Sede", "Prestado", "Recaudado", "Saldo",
            "Int. causado", "Int. recaudado",
            "Mora causada", "Mora recaudada", "Dif. total"
        ],
        filas_sede,
        [75, 85, 85, 85, 90, 90, 90, 90, 85]
    )

    elementos.append(PageBreak())
    agregar_encabezado()

    filas_mes = []
    for item in datos["resumen_mensual"]:
        filas_mes.append([
            item["mes"],
            formato_cop(item["interes_corriente_causado"]),
            formato_cop(item["interes_corriente_recaudado"]),
            formato_cop(item["mora_causada"]),
            formato_cop(item["mora_recaudada"]),
            formato_cop(item["diferencia_interes_corriente"]),
            formato_cop(item["diferencia_mora"]),
            formato_cop(item["total_ingresos"]),
        ])

    tabla_normal(
        "Resumen mensual",
        [
            "Mes", "Int. causado", "Int. recaudado",
            "Mora causada", "Mora recaudada",
            "Dif. interés", "Dif. mora", "Ingresos"
        ],
        filas_mes,
        [90, 95, 95, 95, 95, 95, 95, 95]
    )

    elementos.append(PageBreak())
    agregar_encabezado()

    def filas_tabla_detalle(filas):
        resultado = []
        for f in filas:
            resultado.append([
                f["mes"],
                formato_cop(f["IBAGUE"]),
                formato_cop(f["GIRARDOT"]),
                formato_cop(f["ESPINAL"]),
                formato_cop(f["CRV"]),
                formato_cop(f["TOTAL"]),
            ])
        return resultado

    def agregar_tabla_detalle(titulo, filas, totales):
        cuerpo = filas_tabla_detalle(filas)
        cuerpo.append([
            "TOTALES",
            formato_cop(totales["IBAGUE"]),
            formato_cop(totales["GIRARDOT"]),
            formato_cop(totales["ESPINAL"]),
            formato_cop(totales["CRV"]),
            formato_cop(totales["TOTAL"]),
        ])

        tabla_normal(
            titulo,
            ["Mes", "IBAGUE", "GIRARDOT", "ESPINAL", "CRV", "TOTAL"],
            cuerpo,
            [110, 115, 115, 115, 115, 115]
        )

    agregar_tabla_detalle(
        "Intereses corrientes causados",
        datos["tabla_intereses_causados"],
        datos["totales_intereses_causados"]
    )

    agregar_tabla_detalle(
        "Intereses corrientes recaudados",
        datos["tabla_intereses_recaudados"],
        datos["totales_intereses_recaudados"]
    )

    elementos.append(PageBreak())
    agregar_encabezado()

    agregar_tabla_detalle(
        "Mora causada",
        datos["tabla_mora_causada"],
        datos["totales_mora_causada"]
    )

    agregar_tabla_detalle(
        "Mora recaudada",
        datos["tabla_mora_recaudada"],
        datos["totales_mora_recaudada"]
    )

    elementos.append(PageBreak())
    agregar_encabezado()

    agregar_tabla_detalle(
        "Diferencia intereses corrientes",
        datos["tabla_diferencia_intereses"],
        datos["totales_diferencia_intereses"]
    )

    agregar_tabla_detalle(
        "Diferencia mora",
        datos["tabla_diferencia_mora"],
        datos["totales_diferencia_mora"]
    )

    doc.build(elementos)
    output.seek(0)

    nombre = f"reporte_financiero_{anio_seleccionado}_{sede_seleccionada}.pdf"

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/pdf"
    )

@app.route('/dashboard_gerencial')
def dashboard_gerencial():
    if 'user' not in session:
        return redirect('/login')

    hoy = date.today()
    anio_actual = hoy.year
    mes_actual = hoy.month

    anio_seleccionado = request.args.get('anio', type=int) or anio_actual
    mes_seleccionado = request.args.get('mes', type=int) or mes_actual
    sede_seleccionada = request.args.get('sede', default='TODAS', type=str).strip().upper()

    sedes = ['IBAGUE', 'ESPINAL', 'GIRARDOT', 'CRV']

    if sede_seleccionada == 'TODAS':
        sedes_filtradas = sedes
    else:
        sedes_filtradas = [sede_seleccionada]

    meses_nombres = {
        1: 'ENERO',
        2: 'FEBRERO',
        3: 'MARZO',
        4: 'ABRIL',
        5: 'MAYO',
        6: 'JUNIO',
        7: 'JULIO',
        8: 'AGOSTO',
        9: 'SEPTIEMBRE',
        10: 'OCTUBRE',
        11: 'NOVIEMBRE',
        12: 'DICIEMBRE'
    }

    kpis = {
        'cartera_activa': 0,
        'saldo_total': 0,
        'recaudo_mes': 0,
        'creditos_en_mora': 0,
        'creditos_al_dia': 0,
        'creditos_cancelados': 0,
        'mora_acumulada': 0
    }

    datos_sedes = []

    for sede in sedes_filtradas:
        creditos = Credito.query.filter_by(sede=sede).all()
        total_creditos = len(creditos)
        saldo_actual_sede = round(sum(c.saldo_actual or 0 for c in creditos), 2)

        creditos_en_mora = 0
        creditos_cancelados = 0
        creditos_al_dia = 0
        recaudo_mes_sede = 0
        mora_acumulada_sede = 0

        creditos_ids = [c.id for c in creditos]

        if creditos_ids:
            cuotas = Cuota.query.filter(Cuota.credito_id.in_(creditos_ids)).all()
            cuotas_ids = [q.id for q in cuotas]

            for credito in creditos:
                cuotas_credito = [q for q in cuotas if q.credito_id == credito.id]

                if not cuotas_credito:
                    continue

                if any(q.estado == 'EN MORA' for q in cuotas_credito):
                    creditos_en_mora += 1
                elif all(q.estado in ['PAGADA', 'LIQUIDADA'] for q in cuotas_credito):
                    creditos_cancelados += 1
                else:
                    creditos_al_dia += 1

            mora_acumulada_sede = round(sum(q.interes_mora or 0 for q in cuotas), 2)

            if cuotas_ids:
                pagos_mes_sede_lista = Pago.query.filter(
                    Pago.cuota_id.in_(cuotas_ids),
                    db.extract('year', Pago.fecha) == anio_seleccionado,
                    db.extract('month', Pago.fecha) == mes_seleccionado
                ).all()

                recaudo_mes_sede = round(sum(p.valor or 0 for p in pagos_mes_sede_lista), 2)

        kpis['cartera_activa'] += total_creditos
        kpis['saldo_total'] += saldo_actual_sede
        kpis['recaudo_mes'] += recaudo_mes_sede
        kpis['creditos_en_mora'] += creditos_en_mora
        kpis['creditos_cancelados'] += creditos_cancelados
        kpis['creditos_al_dia'] += creditos_al_dia
        kpis['mora_acumulada'] += mora_acumulada_sede

        datos_sedes.append({
            'sede': sede,
            'total_creditos': total_creditos,
            'saldo_actual': saldo_actual_sede,
            'recaudo_mes': recaudo_mes_sede,
            'en_mora': creditos_en_mora,
            'al_dia': creditos_al_dia,
            'cancelados': creditos_cancelados
        })

    for clave in kpis:
        if isinstance(kpis[clave], float):
            kpis[clave] = round(kpis[clave], 2)

    labels_sedes = [d['sede'] for d in datos_sedes]
    saldos_sedes = [d['saldo_actual'] for d in datos_sedes]
    recaudo_mes_sedes = [d['recaudo_mes'] for d in datos_sedes]
    mora_sedes = [d['en_mora'] for d in datos_sedes]
    al_dia_sedes = [d['al_dia'] for d in datos_sedes]
    cancelados_sedes = [d['cancelados'] for d in datos_sedes]

    return render_template(
        'dashboard_gerencial.html',
        kpis=kpis,
        datos_sedes=datos_sedes,
        labels_sedes=labels_sedes,
        saldos_sedes=saldos_sedes,
        recaudo_mes_sedes=recaudo_mes_sedes,
        mora_sedes=mora_sedes,
        al_dia_sedes=al_dia_sedes,
        cancelados_sedes=cancelados_sedes,
        anio_actual=anio_actual,
        mes_actual=mes_actual,
        anio_seleccionado=anio_seleccionado,
        mes_seleccionado=mes_seleccionado,
        sede_seleccionada=sede_seleccionada,
        anios_disponibles=list(range(2024, anio_actual + 2)),
        meses_disponibles=meses_nombres,
        sedes_disponibles=['TODAS'] + sedes
    )

@app.route('/extracto_credito/<int:credito_id>')
def extracto_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)
    actualizar_mora_credito(credito)
    db.session.commit()

    cuotas = Cuota.query.filter_by(credito_id=credito.id).order_by(Cuota.numero).all()
    cuotas_ids = [c.id for c in cuotas]

    pagos = []
    if cuotas_ids:
        pagos = Pago.query.filter(Pago.cuota_id.in_(cuotas_ids)).order_by(Pago.fecha.desc()).all()

    total_pagado = round(sum(p.valor or 0 for p in pagos), 2)
    total_interes_pagado = round(sum(p.valor_aplicado_interes or 0 for p in pagos), 2)
    total_capital_pagado = round(sum(p.valor_aplicado_capital or 0 for p in pagos), 2)
    total_mora_pagada = round(sum(p.valor_aplicado_mora or 0 for p in pagos), 2)
    total_prepago_capital = round(sum(p.valor_aplicado_prepago_capital or 0 for p in pagos), 2)

    deuda_total_hoy = round(sum((c.total_cobro or 0) for c in cuotas if c.estado in ['PENDIENTE', 'EN MORA', 'ABONO']), 2)

    if all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas) and cuotas:
        estado_credito = 'CANCELADO'
    elif any(c.estado == 'EN MORA' for c in cuotas):
        estado_credito = 'EN MORA'
    else:
        estado_credito = 'AL DÍA'

    return render_template(
        'extracto_credito.html',
        credito=credito,
        cuotas=cuotas,
        pagos=pagos,
        total_pagado=total_pagado,
        total_interes_pagado=total_interes_pagado,
        total_capital_pagado=total_capital_pagado,
        total_mora_pagada=total_mora_pagada,
        total_prepago_capital=total_prepago_capital,
        deuda_total_hoy=deuda_total_hoy,
        estado_credito=estado_credito
    )

@app.route('/ver_recibo_pago/<int:pago_id>')
def ver_recibo_pago(pago_id):
    if 'user' not in session:
        return redirect('/login')

    pago = Pago.query.get_or_404(pago_id)
    cuota = Cuota.query.get_or_404(pago.cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)
    cliente = credito

    mora_aplicada = round(pago.valor_aplicado_mora or 0, 2)

    if mora_aplicada <= 0:
        mora_aplicada = round(pago.mora_generada_al_pago or 0, 2)

    if pago.tipo_pago == 'LIQUIDACION_TOTAL':
        saldo_pendiente_credito = 0

    elif pago.observacion == 'ABONO A CAPITAL':
        saldo_pendiente_credito = round((credito.saldo_actual or 0), 2)

    elif cuota.estado in ['PAGADA', 'LIQUIDADA'] or pago.valor_aplicado_capital > 0:
        saldo_pendiente_credito = round((cuota.saldo_restante or 0), 2)

    else:
        saldo_pendiente_credito = round((credito.saldo_actual or 0), 2)

    return render_template(
        'recibo_pago.html',
        pago=pago,
        cuota=cuota,
        credito=credito,
        cliente=cliente,
        saldo_pendiente_credito=saldo_pendiente_credito,
        mora_aplicada=mora_aplicada
    )

@app.route('/ver_recibo_deuda_fecha/<int:credito_id>')
def ver_recibo_deuda_fecha(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    pagos_param = request.args.get('pagos', '').strip()
    if not pagos_param:
        return redirect(f'/ver_cuotas/{credito.id}')

    ids = []
    for item in pagos_param.split(','):
        item = item.strip()
        if item.isdigit():
            ids.append(int(item))

    pagos = Pago.query.filter(Pago.id.in_(ids)).order_by(Pago.fecha.asc()).all()

    filas = []
    total_pagado = 0

    for pago in pagos:
        cuota = Cuota.query.get(pago.cuota_id)
        if not cuota:
            continue

        filas.append({
            'cuota_numero': cuota.numero,
            'valor_cuota': cuota.valor_cuota,
            'valor_pagado': pago.valor,
            'medio_pago': pago.medio_pago,
            'mora_aplicada': pago.valor_aplicado_mora or 0
        })
        total_pagado += round(pago.valor or 0, 2)

    return render_template(
        'recibo_deuda_fecha.html',
        credito=credito,
        pagos=pagos,
        filas=filas,
        total_pagado=round(total_pagado, 2)
    )

@app.route('/extracto_credito_pdf/<int:credito_id>')
def extracto_credito_pdf(credito_id):
    if 'user' not in session:
        return redirect('/login')

    from io import BytesIO
    from datetime import date
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    def money(valor):
        try:
            return "$ {:,.0f}".format(float(valor or 0)).replace(",", ".")
        except:
            return "$ 0"

    def fecha(valor):
        try:
            return valor.strftime('%d/%m/%Y') if valor else ""
        except:
            return ""

    credito = Credito.query.get_or_404(credito_id)

    actualizar_mora_credito(credito)
    db.session.commit()

    cuotas = Cuota.query.filter_by(
        credito_id=credito.id
    ).order_by(Cuota.numero.asc()).all()

    cuotas_ids = [c.id for c in cuotas]

    pagos = []
    if cuotas_ids:
        pagos = Pago.query.filter(
            Pago.cuota_id.in_(cuotas_ids),
            Pago.tipo_pago != 'ABONO_CAPITAL'
        ).order_by(Pago.fecha.desc()).all()

    abonos_capital = AbonoCapital.query.filter_by(
        credito_id=credito.id
    ).order_by(AbonoCapital.fecha.desc()).all()

    inyecciones = InyeccionCapital.query.filter_by(
        credito_id=credito.id
    ).order_by(InyeccionCapital.numero_cuota.asc()).all()

    cuotas_dict = {c.id: c for c in cuotas}

    total_pagado = round(sum((p.valor or 0) for p in pagos), 2)
    total_interes = round(sum((p.valor_aplicado_interes or 0) for p in pagos), 2)
    total_capital = round(sum((p.valor_aplicado_capital or 0) for p in pagos), 2)
    total_mora_pagada = round(sum((p.valor_aplicado_mora or 0) for p in pagos), 2)
    total_prepago = round(sum((p.valor_aplicado_prepago_capital or 0) for p in pagos), 2)
    total_abonos = round(sum((a.valor or 0) for a in abonos_capital), 2)
    total_inyectado = round(sum((i.valor or 0) for i in inyecciones), 2)

    deuda_total = round(sum((c.total_cobro or 0) for c in cuotas if c.estado in ['PENDIENTE', 'EN MORA', 'ABONO']), 2)
    mora_total = round(sum((c.interes_mora or 0) for c in cuotas), 2)

    estado_credito = 'EN MORA' if any(c.estado == 'EN MORA' for c in cuotas) else 'AL DÍA'

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=28,
        leftMargin=28,
        topMargin=24,
        bottomMargin=24
    )

    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        'TituloExtracto',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#082A4D'),
        alignment=1,
        spaceAfter=4
    )

    normal_center = ParagraphStyle(
        'NormalCenter',
        parent=styles['BodyText'],
        fontSize=8,
        alignment=1,
        leading=10
    )

    seccion_style = ParagraphStyle(
        'Seccion',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#082A4D'),
        spaceBefore=12,
        spaceAfter=6
    )

    elementos = []

    header = Table([
        [
            Paragraph("<b>CRV</b><br/><font size='7'>CONSTRUCCIONES Y<br/>URBANIZACIONES</font>", normal_center),
            Paragraph(
                "<b>EXTRACTO FINANCIERO DEL CRÉDITO</b><br/>"
                "<font size='8'>CONSTRUCCIONES Y URBANIZACIONES S.A.S<br/>"
                "NIT: 901.527.083-2 - IBAGUÉ, TOLIMA</font>",
                titulo_style
            ),
            Paragraph(
                f"<b>Fecha:</b> {date.today().strftime('%d/%m/%Y')}<br/>"
                f"<b>Pagaré:</b> {getattr(credito, 'numero_pagare', None) or getattr(credito, 'pagare', '')}<br/>"
                f"<b>Estado:</b> {estado_credito}",
                styles['BodyText']
            )
        ]
    ], colWidths=[3.2*cm, 10.5*cm, 5.5*cm])

    header.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#E5E7EB')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))

    elementos.append(header)
    elementos.append(Spacer(1, 10))

    datos_cliente = Table([
        ['CLIENTE', credito.cliente, 'CÉDULA', getattr(credito, 'cedula_cliente', '')],
        ['MONTO CRÉDITO', money(getattr(credito, 'monto', 0)), 'INTERÉS', f"{credito.interes}%"],
        ['FECHA CRÉDITO', fecha(getattr(credito, 'fecha_creacion', None)), 'DEUDA A LA FECHA', money(deuda_total)]
    ], colWidths=[3.5*cm, 6.7*cm, 3.5*cm, 5.5*cm])

    datos_cliente.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#082A4D')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#082A4D')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.white),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (3, 0), (3, -1), colors.HexColor('#F8FAFC')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))

    elementos.append(datos_cliente)
    elementos.append(Spacer(1, 12))

    resumen = Table([
        ['TOTAL PAGADO', 'CAPITAL PAGADO', 'INTERÉS PAGADO', 'MORA', 'ABONOS CAPITAL', 'INYECCIONES'],
        [
            money(total_pagado),
            money(total_capital),
            money(total_interes),
            money(mora_total),
            money(total_abonos + total_prepago),
            money(total_inyectado)
        ]
    ], colWidths=[3.2*cm] * 6)

    resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4AF37')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#EAF5FF')),
        ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#E8F5E9')),
        ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#FDECEC')),
        ('BACKGROUND', (4, 1), (4, 1), colors.HexColor('#FFF7E6')),
        ('BACKGROUND', (5, 1), (5, 1), colors.HexColor('#EAF5FF')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#E5E7EB')),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elementos.append(resumen)

    elementos.append(Paragraph("Historial de pagos", seccion_style))

    pagos_data = [[
        'Fecha',
        'N° cuota',
        'Valor pagado',
        'Interés',
        'Capital',
        'Mora',
        'Prepago capital',
        'Medio',
        'Tipo'
    ]]

    for pago in pagos:
        cuota = cuotas_dict.get(pago.cuota_id)
        numero_cuota = cuota.numero if cuota else ''

        pagos_data.append([
            fecha(pago.fecha),
            numero_cuota,
            money(pago.valor),
            money(pago.valor_aplicado_interes),
            money(pago.valor_aplicado_capital),
            money(pago.valor_aplicado_mora),
            money(pago.valor_aplicado_prepago_capital),
            pago.medio_pago or '',
            pago.tipo_pago or 'PAGO'
        ])

    if len(pagos_data) == 1:
        pagos_data.append(['Sin pagos registrados', '', '', '', '', '', '', '', ''])

    tabla_pagos = Table(
        pagos_data,
        repeatRows=1,
        colWidths=[2.1*cm, 1.5*cm, 2.5*cm, 2.2*cm, 2.2*cm, 1.8*cm, 2.5*cm, 2.1*cm, 2.4*cm]
    )

    estilo_pagos = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#082A4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]

    for i in range(1, len(pagos_data)):
        color = colors.HexColor('#FFFFFF') if i % 2 else colors.HexColor('#F8FAFC')
        estilo_pagos.append(('BACKGROUND', (0, i), (-1, i), color))

    tabla_pagos.setStyle(TableStyle(estilo_pagos))
    elementos.append(tabla_pagos)

    elementos.append(Paragraph("Abonos a capital", seccion_style))

    abonos_data = [[
        'Fecha',
        'Valor abonado',
        'Medio de pago',
        'Observación'
    ]]

    for abono in abonos_capital:
        abonos_data.append([
            fecha(abono.fecha),
            money(abono.valor),
            abono.medio_pago or '',
            abono.observacion or 'Abono a capital'
        ])

    if len(abonos_data) == 1:
        abonos_data.append(['Sin abonos a capital registrados', '', '', ''])

    tabla_abonos = Table(
        abonos_data,
        repeatRows=1,
        colWidths=[3*cm, 4*cm, 4*cm, 8.2*cm]
    )

    estilo_abonos = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#082A4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F5E9')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]

    tabla_abonos.setStyle(TableStyle(estilo_abonos))
    elementos.append(tabla_abonos)

    elementos.append(Paragraph("Inyecciones de capital", seccion_style))

    iny_data = [[
        'Fecha',
        'N° cuota',
        'Valor inyectado',
        'Observación'
    ]]

    for iny in inyecciones:
        iny_data.append([
            fecha(iny.fecha),
            iny.numero_cuota,
            money(iny.valor),
            iny.observacion or ''
        ])

    if len(iny_data) == 1:
        iny_data.append(['Sin inyecciones registradas', '', '', ''])

    tabla_iny = Table(
        iny_data,
        repeatRows=1,
        colWidths=[3*cm, 2.5*cm, 4*cm, 9.7*cm]
    )

    estilo_iny = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#082A4D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFF7E6')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]

    tabla_iny.setStyle(TableStyle(estilo_iny))
    elementos.append(tabla_iny)

    elementos.append(Spacer(1, 12))

    nota = Paragraph(
        "<font size='8' color='#475569'>"
        "Este extracto fue generado automáticamente con base en los pagos, abonos, inyecciones de capital y movimientos registrados en el sistema financiero."
        "</font>",
        styles['BodyText']
    )

    elementos.append(nota)

    doc.build(elementos)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=False,
        download_name=f"extracto_credito_{credito.id}.pdf",
        mimetype='application/pdf'
    )

@app.route('/recibo_pago/<int:pago_id>')
def recibo_pago_pdf(pago_id):
    if 'user' not in session:
        return redirect('/login')

    pago = Pago.query.get_or_404(pago_id)
    cuota = Cuota.query.get_or_404(pago.cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)

    buffer = BytesIO()
    from reportlab.pdfgen import canvas
    pdf = canvas.Canvas(buffer, pagesize=letter)

    width, height = letter  # 612 x 792
    centro_x = width / 2    # 306

    gris = colors.HexColor('#e6e6e6')

    def caja(x, y, w, h, texto=None, fill_color=None, bold=False, center=False, font_size=10):
        pdf.setStrokeColor(colors.black)
        pdf.setLineWidth(0.8)

        if fill_color:
            pdf.setFillColor(fill_color)
            pdf.rect(x, y, w, h, fill=1, stroke=1)
        else:
            pdf.setFillColor(colors.white)
            pdf.rect(x, y, w, h, fill=1, stroke=1)

        pdf.setFillColor(colors.black)

        if texto is not None:
            texto = str(texto).strip()
            fuente = "Helvetica-Bold" if bold else "Helvetica"
            tam = font_size

            while tam > 6 and stringWidth(texto, fuente, tam) > (w - 8):
                tam -= 0.5

            pdf.setFont(fuente, tam)

            if center:
                pdf.drawCentredString(x + (w / 2), y + (h / 2) - (tam / 3), texto)
            else:
                pdf.drawString(x + 4, y + (h / 2) - (tam / 3), texto)

    # =========================
    # ENCABEZADO SUPERIOR
    # =========================
    logo_path = os.path.join(app.static_folder, 'logo.png')
    if os.path.exists(logo_path):
        pdf.drawImage(ImageReader(logo_path), 70, 650, width=105, height=70, mask='auto')

    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(centro_x, 705, "CONSTRUCCIONES Y URBANIZACIONES S.A.S")
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(centro_x, 691, "NIT: 901.527.083-2")
    pdf.drawCentredString(centro_x, 677, "AV. AMBALÁ N° 27-136 - PISO 3")
    pdf.drawCentredString(centro_x, 663, "IBAGUÉ - TOLIMA")
    pdf.drawCentredString(centro_x, 649, "TELÉFONO: 311 414 5843")

    # =========================
    # BLOQUES SUPERIORES
    # =========================
    # Todo el bloque útil se centra visualmente entre x=55 y x=557
    # ancho total aprox = 502 pt

    # Cliente
    x1 = 55
    y1 = 555
    label_w = 95
    value_w = 205
    row_h = 26

    datos_cliente = [
        ("RECIBIDO DE", credito.cliente),
        ("DOCUMENTO", credito.cedula_cliente or ""),
        ("DIRECCIÓN", credito.direccion_cliente or ""),
        ("TELÉFONO", credito.telefono_1 or credito.telefono_2 or ""),
        ("CIUDAD", credito.sede or "IBAGUÉ"),
    ]

    y = y1
    for etiqueta, valor in datos_cliente:
        caja(x1, y, label_w, row_h, etiqueta, gris, bold=True, center=True, font_size=8)
        caja(x1 + label_w, y, value_w, row_h, valor, None, bold=False, center=False, font_size=9)
        y -= row_h

    # Fecha y forma de pago
    xr = 355
    caja(xr, 555, 115, 26, "FECHA DE RECIBO", gris, bold=True, center=True, font_size=8)
    caja(xr + 115, 555, 125, 26, fecha_recibo_es(pago.fecha), None, center=True, font_size=9)

    caja(xr, 503, 115, 28, "FORMA DE PAGO", gris, bold=True, center=True, font_size=8)
    caja(xr + 115, 503, 125, 28, pago.medio_pago.upper(), None, center=True, font_size=9)

    entidad_pago = pago.medio_pago.upper()
    if entidad_pago == "EFECTIVO":
        entidad_pago = "CAJA"

    caja(xr, 471, 115, 28, "ENTIDAD DE PAGO", gris, bold=True, center=True, font_size=8)
    caja(xr + 115, 471, 125, 28, entidad_pago, None, center=True, font_size=9)

    # =========================
    # TABLA CENTRAL
    # =========================
    tabla_x = 55
    tabla_y = 340

    caja(tabla_x, tabla_y + 66, 32, 18, "CANT", gris, bold=True, center=True, font_size=8)
    caja(tabla_x + 32, tabla_y + 66, 270, 18, "DESCRIPCIÓN", gris, bold=True, center=True, font_size=8)
    caja(tabla_x + 302, tabla_y + 66, 140, 18, "PAGO", gris, bold=True, center=True, font_size=8)

    descripcion = f"PAGO CUOTA N° {cuota.numero} DE {credito.cuotas} MES: {MESES_ES[cuota.fecha_pago.month]}"

    caja(tabla_x, tabla_y + 33, 32, 33, "1", None, center=True, font_size=10)
    caja(tabla_x + 32, tabla_y + 33, 220, 33, descripcion, None, bold=True, center=True, font_size=9)
    caja(tabla_x + 252, tabla_y + 33, 50, 33, "VALOR", None, bold=True, center=True, font_size=8)
    caja(tabla_x + 302, tabla_y + 33, 140, 33, formato_cop(cuota.valor_cuota), None, bold=True, center=True, font_size=10)

    # =========================
    # RESUMEN
    # =========================
    valor_mora = round(pago.valor_aplicado_mora or 0, 2)
    valor_cuota = round(cuota.valor_cuota or 0, 2)
    valor_exigible = round(valor_cuota + valor_mora, 2)
    valor_pagado = round(pago.valor or 0, 2)
    saldo_pendiente_cuota = round(max(valor_exigible - valor_pagado, 0), 2)

    res_label_x = 335
    res_box_x = 345
    res_box_w = 97

    pdf.setFont("Helvetica-Bold", 10)
    pdf.setFillColor(colors.black)

    pdf.drawRightString(res_label_x, tabla_y + 10, "INTERESES")
    pdf.setFillColor(colors.red)
    pdf.drawRightString(res_box_x + res_box_w - 8, tabla_y + 10, formato_cop(valor_mora))

    pdf.setFillColor(colors.black)
    pdf.drawRightString(res_label_x, tabla_y - 18, "VALOR EXIGIBLE")
    caja(res_box_x, tabla_y - 30, res_box_w, 24, formato_cop(valor_exigible), None, bold=True, center=True, font_size=10)

    pdf.drawRightString(res_label_x, tabla_y - 48, "TOTAL PAGADO")
    caja(res_box_x, tabla_y - 60, res_box_w, 24, formato_cop(valor_pagado), None, bold=True, center=True, font_size=10)

    if saldo_pendiente_cuota > 0:
        pdf.drawRightString(res_label_x, tabla_y - 78, "SALDO PENDIENTE CUOTA")
        caja(res_box_x, tabla_y - 90, res_box_w, 24, formato_cop(saldo_pendiente_cuota), None, bold=True, center=True, font_size=10)
        base_y = 70
    else:
        base_y = 100

    # =========================
    # BLOQUE INFERIOR
    # =========================
    caja(55, base_y + 28, 120, 40, "VALOR RECIBIDO EN LETRA", gris, bold=True, center=True, font_size=8)
    caja(175, base_y + 28, 310, 40, numero_a_letras(valor_pagado), None, center=True, font_size=10)

    observacion = pago.observacion or f"PAGO REGISTRADO EL {fecha_recibo_es(pago.fecha)}"
    caja(55, base_y - 12, 120, 40, "OBSERVACIÓN", gris, bold=True, center=True, font_size=8)
    caja(175, base_y - 12, 310, 40, observacion, None, center=False, font_size=9)

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"recibo_pago_{pago.id}.pdf",
        mimetype="application/pdf"
    )

@app.route('/detalle_pago/<int:pago_id>')
def detalle_pago(pago_id):
    if 'user' not in session:
        return redirect('/login')

    pago = Pago.query.get_or_404(pago_id)
    cuota = Cuota.query.get_or_404(pago.cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)

    return render_template(
        'detalle_pago.html',
        pago=pago,
        cuota=cuota,
        credito=credito
    )

@app.route('/detalle_pago_pdf/<int:pago_id>')
def detalle_pago_pdf(pago_id):
    if 'user' not in session:
        return redirect('/login')

    pago = Pago.query.get_or_404(pago_id)
    cuota = Cuota.query.get_or_404(pago.cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)

    buffer = BytesIO()
    from reportlab.pdfgen import canvas
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    pdf.setFillColor(colors.white)
    pdf.rect(0, 0, width, height, fill=1, stroke=0)

    # Logo
    logo_path = os.path.join(app.static_folder, 'logo.png')
    if os.path.exists(logo_path):
        pdf.drawImage(ImageReader(logo_path), 40, 705, width=90, height=55, mask='auto')

    # Encabezado empresa
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(320, 745, "CONSTRUCCIONES Y URBANIZACIONES S.A.S")
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(320, 730, "NIT: 901.527.083-2")
    pdf.drawCentredString(320, 715, "AV. AMBALÁ N° 27-136 - PISO 3")
    pdf.drawCentredString(320, 700, "IBAGUÉ - TOLIMA")
    pdf.drawCentredString(320, 685, "TELÉFONO: 311 414 5843")

    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(40, 645, "DETALLE DEL PAGO")

    pdf.setFont("Helvetica", 11)
    y = 615
    salto = 20

    datos = [
        ("Cliente", credito.cliente),
        ("Pagaré", credito.numero_pagare),
        ("Cuota", cuota.numero),
        ("Fecha del pago", pago.fecha.strftime('%d/%m/%Y') if pago.fecha else ''),
        ("Medio de pago", pago.medio_pago or ''),
        ("Tipo de pago", pago.tipo_pago or 'No definido'),
        ("Valor pagado", formato_cop(pago.valor)),
        ("Saldo pendiente antes del pago", formato_cop(pago.saldo_pendiente_antes_pago)),
        ("Total exigible al pago", formato_cop(pago.total_exigible_al_pago)),
        ("Días de mora al pago", pago.dias_mora_pagados if pago.dias_mora_pagados is not None else 0),
        ("Mora generada al pago", formato_cop(pago.mora_generada_al_pago)),
        ("Aplicado a interés", formato_cop(pago.valor_aplicado_interes)),
        ("Aplicado a capital", formato_cop(pago.valor_aplicado_capital)),
        ("Aplicado a mora", formato_cop(pago.valor_aplicado_mora)),
        ("Prepago a capital", formato_cop(pago.valor_aplicado_prepago_capital)),
    ]

    for etiqueta, valor in datos:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(40, y, f"{etiqueta}:")
        pdf.setFont("Helvetica", 11)
        pdf.drawString(220, y, str(valor))
        y -= salto

        if y < 100:
            pdf.showPage()
            y = 750

    if pago.observacion:
        y -= 10
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(40, y, "Observación:")
        y -= 18
        pdf.setFont("Helvetica", 11)
        pdf.drawString(40, y, str(pago.observacion))

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"detalle_pago_{pago.id}.pdf",
        mimetype="application/pdf"
    )

@app.route('/eliminar_credito/<int:credito_id>', methods=['POST'])
def eliminar_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)
    sede = credito.sede
    origen = request.form.get('origen', 'activos')

    # Borrar abonos a capital del crédito
    AbonoCapital.query.filter_by(credito_id=credito.id).delete()

    # Borrar inyecciones de capital del crédito
    InyeccionCapital.query.filter_by(credito_id=credito.id).delete()

    # Borrar cambios de tasa del crédito
    CambioTasaInteresCredito.query.filter_by(credito_id=credito.id).delete()

    # Borrar pagos y cuotas manualmente para evitar residuos
    cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

    for cuota in cuotas:
        Pago.query.filter_by(cuota_id=cuota.id).delete()

    Cuota.query.filter_by(credito_id=credito.id).delete()
    try:
        db.session.delete(credito)
        db.session.commit()
        flash('Crédito eliminado correctamente.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar crédito: {str(e)}', 'error')

    if origen == 'cancelados':
        return redirect(url_for('ver_creditos_cancelados', sede=sede))

    return redirect(url_for('ver_creditos', sede=sede))

@app.route('/paz_y_salvo')
def paz_y_salvo():
    if 'user' not in session:
        return redirect('/login')

    q = request.args.get('q', '').strip()

    creditos = Credito.query.order_by(Credito.fecha_creacion.desc()).all()

    resultados = []
    for credito in creditos:
        if not credito_esta_cancelado(credito):
            continue

        texto_busqueda = " ".join([
            str(credito.cliente or ''),
            str(credito.cedula_cliente or ''),
            str(credito.numero_pagare or ''),
            str(credito.sede or '')
        ]).lower()

        if q and q.lower() not in texto_busqueda:
            continue

        resultados.append(credito)

    return render_template(
        'paz_y_salvo.html',
        creditos=resultados,
        q=q
    )

@app.route('/pagare_credito/<int:credito_id>')
def pagare_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    fecha_suscripcion = obtener_fecha_credito_real(credito)
    fecha_inicial_pagos, fecha_vencimiento_final, dia_pago = obtener_primera_y_ultima_cuota(credito)

    valor_pagare = credito.monto_financiado or 0
    valor_pagare_letras = numero_a_letras_es(valor_pagare)

    contexto = {
        'credito': credito,
        'lugar_suscripcion': 'Ibagué, Tolima',
        'fecha_suscripcion': fecha_suscripcion,
        'fecha_suscripcion_texto': fecha_a_ddmmyyyy(fecha_suscripcion),
        'valor_pagare': valor_pagare,
        'valor_pagare_letras': valor_pagare_letras,
        'plazo_meses': credito.cuotas,
        'dia_pago': dia_pago,
        'fecha_inicial_pagos': fecha_inicial_pagos,
        'fecha_vencimiento_final': fecha_vencimiento_final,
        'fecha_inicial_pagos_texto': fecha_a_ddmmyyyy(fecha_inicial_pagos),
        'fecha_vencimiento_final_texto': fecha_a_ddmmyyyy(fecha_vencimiento_final),
        'fecha_carta_texto': formatear_fecha_larga(fecha_suscripcion),
        'cuota_mensual': credito.cuota_mensual or 0,
        'interes_mensual': credito.interes or 0,
        'desde_creacion': request.args.get('desde') == 'crear',
        'volver_url': url_for('dashboard') if request.args.get('desde') == 'crear'
                    else url_for('ver_cuotas', credito_id=credito.id)
    }

    return render_template('pagare_credito.html', **contexto)

def credito_esta_cancelado(credito):
    cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

    if not cuotas:
        return False

    return all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas)

@app.route('/generar_paz_y_salvo/<int:credito_id>')
def generar_paz_y_salvo(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if not credito_esta_cancelado(credito):
        return "Este crédito aún no está cancelado o liquidado."

    fecha_hoy = date.today()
    fecha_credito = credito.fecha_creacion.date() if credito.fecha_creacion else fecha_hoy

    monto_letras = numero_a_letras(credito.monto_financiado or 0)
    monto_numero = formato_cop(credito.monto_financiado or 0)

    fecha_actual = fecha_documento_es(fecha_hoy)
    fecha_credito_larga = fecha_documento_es(fecha_credito)

    return render_template(
        'paz_y_salvo_documento.html',
        credito=credito,
        fecha_actual=fecha_actual,
        fecha_credito_larga=fecha_credito_larga,
        monto_letras=monto_letras,
        monto_numero=monto_numero
    )

@app.route('/plan_pagos/<int:credito_id>')
def plan_pagos_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)
    cuotas = Cuota.query.filter_by(credito_id=credito.id).order_by(Cuota.numero.asc()).all()

    fecha_credito = credito.fecha_creacion.date() if credito.fecha_creacion else date.today()

    return render_template(
        'plan_pagos.html',
        credito=credito,
        cuotas=cuotas,
        fecha_credito=fecha_credito.strftime('%d/%m/%Y')
    )

@app.route('/agregar_sede', methods=['GET', 'POST'])
def agregar_sede():
    if 'user' not in session:
        return redirect('/login')

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip().upper()

        if not nombre:
            flash('Debes ingresar el nombre de la sede.', 'error')
            return render_template('agregar_sede.html')

        existente = Sede.query.filter_by(nombre=nombre).first()
        if existente:
            flash('Esa sede ya existe.', 'error')
            return render_template('agregar_sede.html')

        try:
            nueva_sede = Sede(nombre=nombre, activa=True)
            db.session.add(nueva_sede)
            db.session.commit()
            flash('Sede agregada correctamente.', 'success')
            return redirect('/dashboard')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar sede: {str(e)}', 'error')

    return render_template('agregar_sede.html')

@app.route('/inyeccion_capital/<int:credito_id>', methods=['GET', 'POST'])
def inyeccion_capital(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if credito.tipo_cuota != 'VARIABLE':
        flash("Este crédito no es de cuota variable.", "error")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    if not credito.permite_inyeccion_capital:
        flash("Este crédito no permite inyección de capital.", "error")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    if request.method == 'POST':
        numero_cuota = int(request.form['numero_cuota'])
        fecha = datetime.strptime(request.form['fecha'], '%Y-%m-%d')
        valor = limpiar_valor_moneda(request.form['valor'])
        observacion = request.form.get('observacion', '').strip()

        # No permitir recalcular cuotas futuras que ya tengan pagos
        cuotas_futuras_ids = [
            c.id for c in Cuota.query.filter(
                Cuota.credito_id == credito.id,
                Cuota.numero >= numero_cuota
            ).all()
        ]

        if cuotas_futuras_ids:
            pagos_futuros = Pago.query.filter(Pago.cuota_id.in_(cuotas_futuras_ids)).count()

            if pagos_futuros > 0:
                flash(
                    "No se puede aplicar esta inyección porque existen pagos registrados "
                    "en la cuota seleccionada o en cuotas posteriores.",
                    "error"
                )
                return redirect(url_for('ver_cuotas', credito_id=credito.id))

        nueva = InyeccionCapital(
            credito_id=credito.id,
            numero_cuota=numero_cuota,
            fecha=fecha,
            valor=valor,
            observacion=observacion
        )

        db.session.add(nueva)
        db.session.flush()

        # Tomar saldo desde la cuota anterior para NO tocar historia ya pagada
        cuota_anterior = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero < numero_cuota
        ).order_by(Cuota.numero.desc()).first()

        if cuota_anterior:
            saldo = cuota_anterior.saldo_restante
        else:
            saldo = credito.monto_financiado

        # Borrar solo cuotas desde la inyección hacia adelante
        Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero >= numero_cuota
        ).delete()
        db.session.commit()

        dia_original = credito.fecha_creacion.day

        for numero in range(numero_cuota, credito.cuotas + 1):
            fecha_pago = sumar_meses(
                credito.fecha_creacion,
                numero - 1,
                dia_fijo=dia_original
            )

            tasa_mes = obtener_tasa_interes_variable(
                fecha_pago.year,
                credito.interes
            )

            inyecciones = InyeccionCapital.query.filter_by(
                credito_id=credito.id,
                numero_cuota=numero
            ).all()

            adicion_capital = round(sum(i.valor or 0 for i in inyecciones), 2)

            saldo_inicial = round(saldo + adicion_capital, 2)

            interes_mes = round(saldo_inicial * (tasa_mes / 100), 2)

            cuotas_restantes = credito.cuotas - numero + 1
            capital = round(saldo_inicial / max(cuotas_restantes, 1), 2)

            valor_cuota = round(capital + interes_mes, 2)

            saldo = round(saldo_inicial - capital, 2)

            if saldo < 0:
                capital = round(capital + saldo, 2)
                saldo = 0

            nueva_cuota = Cuota(
                credito_id=credito.id,
                numero=numero,
                fecha_pago=fecha_pago,
                valor_cuota=valor_cuota,
                saldo_inicial=saldo_inicial,
                capital=capital,
                interes=interes_mes,
                saldo_restante=saldo,
                saldo_pendiente=valor_cuota,
                tasa_mora_mensual_cuota=0,
                porcentaje_mora_aplicado=0,
                dias_mora=0,
                interes_mora=0,
                total_cobro=valor_cuota,
                estado='PENDIENTE'
            )

            db.session.add(nueva_cuota)

        db.session.commit()

        return redirect(url_for(
            'ver_recibo_inyeccion_capital',
            inyeccion_id=nueva.id
        ))

    inyecciones = InyeccionCapital.query.filter_by(
        credito_id=credito.id
    ).order_by(InyeccionCapital.numero_cuota.asc()).all()

    return render_template(
        'inyeccion_capital.html',
        credito=credito,
        inyecciones=inyecciones
    )

@app.route('/simular_inyeccion/<int:credito_id>', methods=['GET', 'POST'])
def simular_inyeccion(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    cuotas_simuladas = []

    if request.method == 'POST':
        numero_cuota = int(request.form['numero_cuota'])
        fecha = datetime.strptime(request.form['fecha'], '%Y-%m-%d')
        valor = limpiar_valor_moneda(request.form['valor'])

        # Buscar saldo base
        cuota_anterior = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero < numero_cuota
        ).order_by(Cuota.numero.desc()).first()

        if cuota_anterior:
            saldo = cuota_anterior.saldo_restante
        else:
            saldo = credito.monto_financiado

        dia_original = credito.fecha_creacion.day

        for numero in range(numero_cuota, credito.cuotas + 1):

            fecha_pago = sumar_meses(
                credito.fecha_creacion,
                numero - 1,
                dia_fijo=dia_original
            )

            tasa_mes = obtener_tasa_interes_variable(
                fecha_pago.year,
                credito.interes
            )

            # Inyección SOLO en la cuota elegida
            adicion_capital = valor if numero == numero_cuota else 0

            saldo_inicial = round(saldo + adicion_capital, 2)

            interes_mes = round(saldo_inicial * (tasa_mes / 100), 2)

            cuotas_restantes = credito.cuotas - numero + 1
            capital = round(saldo_inicial / max(cuotas_restantes, 1), 2)

            valor_cuota = round(capital + interes_mes, 2)

            saldo = round(saldo_inicial - capital, 2)

            cuotas_simuladas.append({
                'numero': numero,
                'fecha': fecha_pago,
                'saldo_inicial': saldo_inicial,
                'adicion': adicion_capital,
                'interes': interes_mes,
                'capital': capital,
                'cuota': valor_cuota,
                'saldo_final': saldo
            })

    return render_template(
        'simular_inyeccion.html',
        credito=credito,
        cuotas=cuotas_simuladas
    )

@app.route('/cambiar_tasa_interes_credito/<int:credito_id>', methods=['GET', 'POST'])
def cambiar_tasa_interes_credito(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if credito.tipo_cuota != 'VARIABLE' or credito.tipo_interes != 'VARIABLE':
        flash("Este crédito no permite cambio de tasa variable.", "error")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    if request.method == 'POST':
        numero_cuota = int(request.form['numero_cuota'])
        fecha_cambio = datetime.strptime(request.form['fecha_cambio'], '%Y-%m-%d')
        tasa_nueva = float(request.form['tasa_nueva'])
        observacion = request.form.get('observacion', '').strip()

        cuotas_futuras_ids = [
            c.id for c in Cuota.query.filter(
                Cuota.credito_id == credito.id,
                Cuota.numero >= numero_cuota
            ).all()
        ]

        if cuotas_futuras_ids:
            pagos_futuros = Pago.query.filter(Pago.cuota_id.in_(cuotas_futuras_ids)).count()

            if pagos_futuros > 0:
                flash(
                    "No se puede cambiar la tasa porque existen pagos registrados "
                    "en la cuota seleccionada o en cuotas posteriores.",
                    "error"
                )
                return redirect(url_for('ver_cuotas', credito_id=credito.id))

        cuota_anterior = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero < numero_cuota
        ).order_by(Cuota.numero.desc()).first()

        if cuota_anterior:
            saldo = cuota_anterior.saldo_restante
            tasa_anterior = cuota_anterior.interes / cuota_anterior.saldo_inicial * 100 if cuota_anterior.saldo_inicial else credito.interes
        else:
            saldo = credito.monto_financiado
            tasa_anterior = credito.interes

        cambio = CambioTasaInteresCredito(
            credito_id=credito.id,
            numero_cuota=numero_cuota,
            fecha_cambio=fecha_cambio,
            tasa_anterior=tasa_anterior,
            tasa_nueva=tasa_nueva,
            observacion=observacion
        )

        db.session.add(cambio)
        db.session.commit()

        Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero >= numero_cuota
        ).delete()
        db.session.commit()

        dia_original = credito.fecha_creacion.day

        for numero in range(numero_cuota, credito.cuotas + 1):
            fecha_pago = sumar_meses(
                credito.fecha_creacion,
                numero - 1,
                dia_fijo=dia_original
            )

            inyecciones = InyeccionCapital.query.filter_by(
                credito_id=credito.id,
                numero_cuota=numero
            ).all()

            adicion_capital = round(sum(i.valor or 0 for i in inyecciones), 2)

            saldo_inicial = round(saldo + adicion_capital, 2)
            interes_mes = round(saldo_inicial * (tasa_nueva / 100), 2)

            cuotas_restantes = credito.cuotas - numero + 1
            capital = round(saldo_inicial / max(cuotas_restantes, 1), 2)

            valor_cuota = round(capital + interes_mes, 2)
            saldo = round(saldo_inicial - capital, 2)

            if saldo < 0:
                capital = round(capital + saldo, 2)
                saldo = 0

            nueva_cuota = Cuota(
                credito_id=credito.id,
                numero=numero,
                fecha_pago=fecha_pago,
                valor_cuota=valor_cuota,
                saldo_inicial=saldo_inicial,
                capital=capital,
                interes=interes_mes,
                saldo_restante=saldo,
                saldo_pendiente=valor_cuota,
                tasa_mora_mensual_cuota=0,
                porcentaje_mora_aplicado=0,
                dias_mora=0,
                interes_mora=0,
                total_cobro=valor_cuota,
                estado='PENDIENTE'
            )

            db.session.add(nueva_cuota)

        credito.saldo_actual = saldo
        db.session.commit()

        flash("Tasa de interés cambiada correctamente.", "success")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    historial = CambioTasaInteresCredito.query.filter_by(
        credito_id=credito.id
    ).order_by(CambioTasaInteresCredito.numero_cuota.asc()).all()

    return render_template(
        'cambiar_tasa_interes_credito.html',
        credito=credito,
        historial=historial
    )

@app.route('/reversar_pago/<int:pago_id>', methods=['POST'])
def reversar_pago(pago_id):
    if 'user' not in session:
        return redirect('/login')

    pago = Pago.query.get_or_404(pago_id)

    if pago.reversado:
        return redirect(request.referrer or '/dashboard')

    cuota = Cuota.query.get_or_404(pago.cuota_id)
    credito = Credito.query.get_or_404(cuota.credito_id)

    motivo = request.form.get('motivo_reversion', '').strip()

    if not motivo:
        flash("Debes escribir el motivo de la reversión.", "error")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    # Marcar pago como reversado, NO borrarlo
    pago.activo = False
    pago.reversado = True
    pago.motivo_reversion = motivo
    pago.fecha_reversion = datetime.now()

    db.session.flush()

    # Recalcular desde la cuota ANTERIOR para reconstruir la tabla
    numero_base = max((cuota.numero or 1) - 1, 0)

    cuota_base = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.numero == numero_base
    ).first()

    if cuota_base:
        fecha_base = (
            cuota_base.fecha_pago.date()
            if isinstance(cuota_base.fecha_pago, datetime)
            else cuota_base.fecha_pago
        )
    else:
        fecha_base = (
            credito.fecha_creacion.date()
            if isinstance(credito.fecha_creacion, datetime)
            else credito.fecha_creacion
        )

    if credito.tipo_cuota == 'VARIABLE' and credito.tipo_interes == 'VARIABLE':
        recalcular_cuotas_variables_pendientes(
            credito=credito,
            cuota_actual_numero=numero_base,
            fecha_base=fecha_base
        )
    else:
        recalcular_cuotas_pendientes(
            credito=credito,
            cuota_actual_numero=numero_base,
            fecha_base=fecha_base
        )

    db.session.commit()

    actualizar_mora_credito(credito, date.today())
    db.session.commit()

    flash("Pago reversado correctamente.", "success")
    return redirect(url_for('ver_cuotas', credito_id=credito.id))

@app.route('/historial_reversiones/<int:credito_id>')
def historial_reversiones(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    pagos_reversados = (
        db.session.query(Pago, Cuota)
        .join(Cuota, Pago.cuota_id == Cuota.id)
        .filter(
            Cuota.credito_id == credito.id,
            Pago.reversado == True
        )
        .order_by(Pago.fecha_reversion.desc())
        .all()
    )

    abonos_reversados = AbonoCapital.query.filter_by(
        credito_id=credito.id,
        reversado=True
    ).order_by(AbonoCapital.fecha_reversion.desc()).all()

    return render_template(
        'historial_reversiones.html',
        credito=credito,
        pagos_reversados=pagos_reversados,
        abonos_reversados=abonos_reversados
    )
    
@app.route('/fix_db')
def fix_db():
    try:
        from sqlalchemy import text

        db.session.execute(text("ALTER TABLE pago ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE"))
        db.session.execute(text("ALTER TABLE pago ADD COLUMN IF NOT EXISTS reversado BOOLEAN DEFAULT FALSE"))
        db.session.execute(text("ALTER TABLE pago ADD COLUMN IF NOT EXISTS motivo_reversion VARCHAR(255)"))
        db.session.execute(text("ALTER TABLE pago ADD COLUMN IF NOT EXISTS fecha_reversion TIMESTAMP"))

        db.session.commit()
        return "Columnas creadas correctamente"
        
    except Exception as e:
        return f"Error: {str(e)}"


@app.route('/recalcular_variable_abonos/<int:credito_id>')
def recalcular_variable_abonos(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if credito.tipo_cuota != 'VARIABLE':
        return "Este recalculo solo aplica para créditos de cuota variable"

    ultima_pagada = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.estado == 'PAGADA'
    ).order_by(Cuota.numero.desc()).first()

    if not ultima_pagada:
        return "No se encontró una cuota pagada como punto de partida"

    recalcular_cuotas_variables_pendientes(
        credito=credito,
        cuota_actual_numero=ultima_pagada.numero,
        fecha_base=ultima_pagada.fecha_pago
    )

    actualizar_mora_credito(credito, date.today())

    db.session.commit()

    return redirect(url_for('ver_cuotas', credito_id=credito.id))

@app.route('/ver_recibo_abono_capital/<int:abono_id>')
def ver_recibo_abono_capital(abono_id):
    if 'user' not in session:
        return redirect('/login')

    abono = AbonoCapital.query.get_or_404(abono_id)
    credito = Credito.query.get_or_404(abono.credito_id)

    return render_template(
        'recibo_abono_capital.html',
        abono=abono,
        credito=credito,
        cliente=credito
    )

@app.route('/ver_recibo_inyeccion_capital/<int:inyeccion_id>')
def ver_recibo_inyeccion_capital(inyeccion_id):
    if 'user' not in session:
        return redirect('/login')

    inyeccion = InyeccionCapital.query.get_or_404(inyeccion_id)
    credito = Credito.query.get_or_404(inyeccion.credito_id)

    return render_template(
        'recibo_inyeccion_capital.html',
        inyeccion=inyeccion,
        credito=credito
    )

@app.route('/ver_pagare/<int:credito_id>')
def ver_pagare(credito_id):
    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    cuotas = Cuota.query.filter_by(
        credito_id=credito.id
    ).order_by(Cuota.numero.asc()).all()

    primera_cuota = cuotas[0] if cuotas else None
    ultima_cuota = cuotas[-1] if cuotas else None

    valor_pagare = credito.monto_financiado or ((credito.monto or 0) - (credito.abono_inicial or 0))
    plazo_meses = credito.cuotas or len(cuotas)

    fecha_suscripcion = credito.fecha_creacion.date() if isinstance(credito.fecha_creacion, datetime) else credito.fecha_creacion

    fecha_inicial_pagos = primera_cuota.fecha_pago if primera_cuota else credito.fecha_creacion
    fecha_vencimiento_final = ultima_cuota.fecha_pago if ultima_cuota else credito.fecha_creacion

    if isinstance(fecha_inicial_pagos, datetime):
        fecha_inicial_pagos = fecha_inicial_pagos.date()

    if isinstance(fecha_vencimiento_final, datetime):
        fecha_vencimiento_final = fecha_vencimiento_final.date()

    cuota_mensual = primera_cuota.valor_cuota if primera_cuota else 0
    interes_mensual = credito.interes or 0
    dia_pago = fecha_inicial_pagos.day if fecha_inicial_pagos else ''

    try:
        valor_pagare_letras = numero_a_letras(int(valor_pagare))
    except:
        valor_pagare_letras = ""

    return render_template(
        'pagare_credito.html',
        volver_url=url_for('ver_cuotas', credito_id=credito.id),
        credito=credito,
        desde_creacion=False,
        lugar_suscripcion="Ibagué - Tolima",
        fecha_suscripcion_texto=fecha_suscripcion.strftime('%d/%m/%Y') if fecha_suscripcion else '',
        fecha_carta_texto=fecha_suscripcion.strftime('%d/%m/%Y') if fecha_suscripcion else '',
        valor_pagare=valor_pagare,
        valor_pagare_letras=valor_pagare_letras,
        plazo_meses=plazo_meses,
        dia_pago=dia_pago,
        fecha_inicial_pagos_texto=fecha_inicial_pagos.strftime('%d/%m/%Y') if fecha_inicial_pagos else '',
        fecha_vencimiento_final_texto=fecha_vencimiento_final.strftime('%d/%m/%Y') if fecha_vencimiento_final else '',
        cuota_mensual=cuota_mensual,
        interes_mensual=interes_mensual
    )

@app.route('/reversar_abono_capital/<int:abono_id>', methods=['POST'])
def reversar_abono_capital(abono_id):
    if 'user' not in session:
        return redirect('/login')

    abono = AbonoCapital.query.get_or_404(abono_id)
    credito = Credito.query.get_or_404(abono.credito_id)

    motivo = request.form.get('motivo_reversion', '').strip()

    if not motivo:
        flash("Debes escribir el motivo de la reversión.", "error")
        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    fecha_abono = abono.fecha.date() if isinstance(abono.fecha, datetime) else abono.fecha

    cuota_afectada = Cuota.query.filter(
        Cuota.credito_id == credito.id,
        Cuota.fecha_pago >= fecha_abono
    ).order_by(Cuota.numero.asc()).first()

    abono.activo = False
    abono.reversado = True
    abono.motivo_reversion = motivo
    abono.fecha_reversion = datetime.now()

    credito.saldo_actual = round((credito.saldo_actual or 0) + (abono.valor or 0), 2)

    db.session.commit()

    if cuota_afectada:

        numero_base = max((cuota_afectada.numero or 1) - 1, 0)

        cuota_base = Cuota.query.filter(
            Cuota.credito_id == credito.id,
            Cuota.numero == numero_base
        ).first()

        if cuota_base:
            fecha_base = cuota_base.fecha_pago
        else:
            fecha_base = credito.fecha_creacion

        if credito.tipo_cuota == 'VARIABLE' and credito.tipo_interes == 'VARIABLE':
            recalcular_cuotas_variables_pendientes(
                credito=credito,
                cuota_actual_numero=numero_base,
                fecha_base=fecha_base
            )
        else:
            recalcular_cuotas_pendientes(
                credito=credito,
                cuota_actual_numero=numero_base,
                fecha_base=fecha_base
            )

    db.session.commit()

    flash("Abono a capital reversado correctamente.", "success")
    return redirect(url_for('ver_cuotas', credito_id=credito.id))

@app.route('/editar_cliente_credito/<int:credito_id>', methods=['GET', 'POST'])
def editar_cliente_credito(credito_id):

    if 'user' not in session:
        return redirect('/login')

    credito = Credito.query.get_or_404(credito_id)

    if request.method == 'POST':

        credito.cliente = request.form.get('cliente')
        credito.telefono_1 = request.form.get('telefono_1')
        credito.telefono_2 = request.form.get('telefono_2')
        credito.direccion_cliente = request.form.get('direccion_cliente')
        credito.correo_cliente = request.form.get('correo_cliente')

        db.session.commit()

        flash('Datos actualizados correctamente', 'success')

        return redirect(url_for('ver_cuotas', credito_id=credito.id))

    return render_template(
        'editar_cliente_credito.html',
        credito=credito
    )

@app.route('/exportar_clientes_sede/<sede>')
def exportar_clientes_sede(sede):
    if 'user' not in session:
        return redirect('/login')

    creditos = Credito.query.filter_by(sede=sede).order_by(Credito.cliente.asc()).all()

    filas = []

    for credito in creditos:
        cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

        if all(c.estado in ['PAGADA', 'LIQUIDADA'] for c in cuotas):
            continue

        filas.append({
            'Cliente': credito.cliente,
            'Cédula': credito.cedula_cliente,
            'Teléfono 1': credito.telefono_1,
            'Teléfono 2': credito.telefono_2,
            'Correo': credito.correo_cliente,
            'Dirección': credito.direccion_cliente,
            'Pagaré': credito.numero_pagare,
            'Sede': credito.sede
        })

    df = pd.DataFrame(filas)

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Clientes activos')
    output.seek(0)

    return send_file(
        output,
        download_name=f'clientes_activos_{sede}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/exportar_clientes_mora_sede/<sede>')
def exportar_clientes_mora_sede(sede):
    if 'user' not in session:
        return redirect('/login')

    sede = sede.strip().upper()
    hoy = date.today()

    creditos = Credito.query.filter_by(sede=sede).order_by(Credito.cliente.asc()).all()

    filas = []

    for credito in creditos:
        actualizar_mora_credito(credito, hoy)

        cuotas = Cuota.query.filter_by(credito_id=credito.id).all()

        cuotas_mora = [
            c for c in cuotas
            if c.estado == 'EN MORA'
        ]

        if not cuotas_mora:
            continue

        deuda_fecha = round(sum(
            (c.saldo_pendiente or 0) + (c.interes_mora or 0)
            for c in cuotas
            if c.estado in ['PENDIENTE', 'EN MORA', 'ABONO']
            and (
                (c.fecha_pago.date() if isinstance(c.fecha_pago, datetime) else c.fecha_pago) <= hoy
            )
        ), 2)

        mora_total = round(sum(
            c.interes_mora or 0
            for c in cuotas_mora
        ), 2)

        cuotas_vencidas = len(cuotas_mora)

        filas.append({
            'Cliente': credito.cliente,
            'Cédula': credito.cedula_cliente,
            'Teléfono 1': credito.telefono_1,
            'Teléfono 2': credito.telefono_2,
            'Correo': credito.correo_cliente,
            'Dirección': credito.direccion_cliente,
            'Pagaré': credito.numero_pagare,
            'Sede': credito.sede,
            'Cuotas en mora': cuotas_vencidas,
            'Mora total': mora_total,
            'Deuda a la fecha': deuda_fecha
        })

    db.session.commit()

    df = pd.DataFrame(filas)

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Clientes en mora')
    output.seek(0)

    return send_file(
        output,
        download_name=f'clientes_en_mora_{sede}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == "__main__":
    app.run(debug=True)

